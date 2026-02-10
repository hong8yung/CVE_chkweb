from __future__ import annotations

import argparse
import json
import math
import re
import time
from io import BytesIO
from datetime import datetime, timedelta
from html import escape
from urllib.parse import urlencode

import psycopg2
from flask import Flask, jsonify, redirect, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from psycopg2.extras import Json

from classification import IMPACT_TYPE_OPTIONS
from nvd_fetch import fetch_cves_from_db, fetch_incremental_checkpoint
from settings import Settings, load_settings

app = Flask(__name__)

COUNT_CACHE_TTL_SECONDS = 120
COUNT_CACHE_MAX_ENTRIES = 200
_count_cache: dict[str, tuple[int, float]] = {}
VALID_USER_PROFILES = {"hq", "jaehwa"}
VALID_SORT_KEYS = {"cvss_desc", "cvss_asc", "last_modified_desc", "last_modified_asc"}
DEFAULT_PROFILE_SETTINGS: dict[str, object] = {
    "vendor": "",
    "product": "",
    "keyword": "",
    "cpe_objects_catalog": [],
    "min_cvss": 0.0,
    "limit": 50,
    "impact_type": [],
    "cpe_missing_only": False,
    "sort_key": "cvss_desc",
    "last_modified_lookback_days": 7,
    "daily_review_window_days": 1,
    "daily_review_limit": 300,
}
_profile_settings_table_ready = False
_review_status_table_ready = False
_profile_presets_table_ready = False
_last_bulk_action_cache: dict[str, dict[str, object]] = {}


def _normalize_user_profile(raw_value: str | None) -> str:
    profile = (raw_value or "hq").strip().lower()
    return profile if profile in VALID_USER_PROFILES else "hq"


def _to_bool(value: object) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _sanitize_profile_settings(raw_settings: dict[str, object] | None) -> dict[str, object]:
    clean = dict(DEFAULT_PROFILE_SETTINGS)
    if not isinstance(raw_settings, dict):
        return clean

    clean["vendor"] = str(raw_settings.get("vendor", "")).strip()
    clean["product"] = str(raw_settings.get("product", "")).strip()
    clean["keyword"] = str(raw_settings.get("keyword", "")).strip()
    cpe_catalog_value = raw_settings.get("cpe_objects_catalog", [])
    if isinstance(cpe_catalog_value, str):
        cpe_catalog_candidates = [value.strip() for value in cpe_catalog_value.replace(",", "\n").splitlines()]
    elif isinstance(cpe_catalog_value, list):
        cpe_catalog_candidates = [str(value).strip() for value in cpe_catalog_value]
    else:
        cpe_catalog_candidates = []
    clean_catalog: list[str] = []
    seen_catalog: set[str] = set()
    for value in cpe_catalog_candidates:
        parts = [part.strip().lower() for part in value.split(":")]
        if len(parts) < 2:
            continue
        if not parts[0] or not parts[1]:
            continue
        normalized = ":".join(parts[:3]) if len(parts) >= 3 and parts[2] else f"{parts[0]}:{parts[1]}"
        if normalized in seen_catalog:
            continue
        seen_catalog.add(normalized)
        clean_catalog.append(normalized)
    clean["cpe_objects_catalog"] = clean_catalog

    try:
        clean["min_cvss"] = max(0.0, min(float(raw_settings.get("min_cvss", 0.0)), 10.0))
    except (TypeError, ValueError):
        clean["min_cvss"] = 0.0

    try:
        clean["limit"] = max(1, min(int(raw_settings.get("limit", 50)), 500))
    except (TypeError, ValueError):
        clean["limit"] = 50

    sort_key = str(raw_settings.get("sort_key", DEFAULT_PROFILE_SETTINGS["sort_key"])).strip().lower()
    clean["sort_key"] = sort_key if sort_key in VALID_SORT_KEYS else DEFAULT_PROFILE_SETTINGS["sort_key"]

    try:
        clean["last_modified_lookback_days"] = max(1, min(int(raw_settings.get("last_modified_lookback_days", 7)), 365))
    except (TypeError, ValueError):
        clean["last_modified_lookback_days"] = 7
    try:
        clean["daily_review_window_days"] = max(1, min(int(raw_settings.get("daily_review_window_days", 1)), 30))
    except (TypeError, ValueError):
        clean["daily_review_window_days"] = 1
    try:
        clean["daily_review_limit"] = max(1, min(int(raw_settings.get("daily_review_limit", 300)), 1000))
    except (TypeError, ValueError):
        clean["daily_review_limit"] = 300

    clean["cpe_missing_only"] = _to_bool(raw_settings.get("cpe_missing_only", False))

    impact_values = raw_settings.get("impact_type", [])
    if isinstance(impact_values, str):
        impact_candidates = [value.strip() for value in impact_values.split(",")]
    elif isinstance(impact_values, list):
        impact_candidates = [str(value).strip() for value in impact_values]
    else:
        impact_candidates = []

    deduped_impacts: list[str] = []
    for value in impact_candidates:
        if not value or value not in IMPACT_TYPE_OPTIONS or value in deduped_impacts:
            continue
        deduped_impacts.append(value)
    clean["impact_type"] = deduped_impacts
    return clean


def _connect_db(settings_obj: Settings) -> psycopg2.extensions.connection:
    return psycopg2.connect(
        host=settings_obj.db_host,
        port=settings_obj.db_port,
        dbname=settings_obj.db_name,
        user=settings_obj.db_user,
        password=settings_obj.db_password,
    )


def _ensure_profile_settings_table(settings_obj: Settings) -> None:
    global _profile_settings_table_ready
    if _profile_settings_table_ready:
        return
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS user_profile_settings (
                      profile_key   text PRIMARY KEY,
                      settings_json jsonb NOT NULL,
                      updated_at    timestamptz NOT NULL DEFAULT now()
                    )
                    """
                )
    finally:
        conn.close()
    _profile_settings_table_ready = True


def _ensure_review_status_table(settings_obj: Settings) -> None:
    global _review_status_table_ready
    if _review_status_table_ready:
        return
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS daily_review_item (
                      profile_key  text NOT NULL,
                      review_date  date NOT NULL,
                      cve_id       text NOT NULL REFERENCES cve (id) ON DELETE CASCADE,
                      status       text NOT NULL DEFAULT 'pending',
                      note         text NOT NULL DEFAULT '',
                      reviewed_at  timestamptz,
                      updated_at   timestamptz NOT NULL DEFAULT now(),
                      PRIMARY KEY (profile_key, review_date, cve_id)
                    )
                    """
                )
    finally:
        conn.close()
    _review_status_table_ready = True


def fetch_profile_settings(settings_obj: Settings, profile: str) -> dict[str, object]:
    normalized_profile = _normalize_user_profile(profile)
    _ensure_profile_settings_table(settings_obj)
    conn = _connect_db(settings_obj)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT settings_json FROM user_profile_settings WHERE profile_key = %s",
                (normalized_profile,),
            )
            row = cur.fetchone()
    finally:
        conn.close()
    if not row:
        return dict(DEFAULT_PROFILE_SETTINGS)
    raw_value = row[0]
    if isinstance(raw_value, dict):
        return _sanitize_profile_settings(raw_value)
    if isinstance(raw_value, str):
        try:
            return _sanitize_profile_settings(json.loads(raw_value))
        except json.JSONDecodeError:
            return dict(DEFAULT_PROFILE_SETTINGS)
    return dict(DEFAULT_PROFILE_SETTINGS)


def upsert_profile_settings(settings_obj: Settings, profile: str, payload: dict[str, object]) -> dict[str, object]:
    normalized_profile = _normalize_user_profile(profile)
    sanitized_payload = _sanitize_profile_settings(payload)
    _ensure_profile_settings_table(settings_obj)
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO user_profile_settings (profile_key, settings_json, updated_at)
                    VALUES (%s, %s, now())
                    ON CONFLICT (profile_key)
                    DO UPDATE SET
                      settings_json = EXCLUDED.settings_json,
                      updated_at = now()
                    """,
                    (normalized_profile, Json(sanitized_payload)),
                )
    finally:
        conn.close()
    return sanitized_payload


def fetch_daily_review_map(
    settings_obj: Settings,
    profile: str,
    review_date: str,
) -> dict[str, dict[str, str]]:
    _ensure_review_status_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    conn = _connect_db(settings_obj)
    result: dict[str, dict[str, str]] = {}
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT cve_id, status, note
                FROM daily_review_item
                WHERE profile_key = %s
                  AND review_date = %s::date
                """,
                (normalized_profile, review_date),
            )
            rows = cur.fetchall()
        for cve_id, status, note in rows:
            result[str(cve_id)] = {"status": str(status), "note": str(note or "")}
    finally:
        conn.close()
    return result


def upsert_daily_review_item(
    settings_obj: Settings,
    profile: str,
    review_date: str,
    cve_id: str,
    status: str,
    note: str,
) -> None:
    _ensure_review_status_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    status_value = status if status in {"pending", "reviewed", "ignored"} else "pending"
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO daily_review_item (profile_key, review_date, cve_id, status, note, reviewed_at, updated_at)
                    VALUES (%s, %s::date, %s, %s, %s, CASE WHEN %s = 'pending' THEN NULL ELSE now() END, now())
                    ON CONFLICT (profile_key, review_date, cve_id)
                    DO UPDATE SET
                      status = EXCLUDED.status,
                      note = EXCLUDED.note,
                      reviewed_at = CASE WHEN EXCLUDED.status = 'pending' THEN NULL ELSE now() END,
                      updated_at = now()
                    """,
                    (normalized_profile, review_date, cve_id, status_value, note[:500], status_value),
                )
    finally:
        conn.close()


def _ensure_profile_presets_table(settings_obj: Settings) -> None:
    global _profile_presets_table_ready
    if _profile_presets_table_ready:
        return
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS user_profile_preset (
                      profile_key  text NOT NULL,
                      preset_name  text NOT NULL,
                      rule_json    jsonb NOT NULL,
                      is_enabled   boolean NOT NULL DEFAULT true,
                      updated_at   timestamptz NOT NULL DEFAULT now(),
                      PRIMARY KEY (profile_key, preset_name)
                    )
                    """
                )
    finally:
        conn.close()
    _profile_presets_table_ready = True


def fetch_profile_presets(settings_obj: Settings, profile: str) -> list[dict[str, object]]:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    conn = _connect_db(settings_obj)
    rows: list[tuple[object, object, object, object]] = []
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT preset_name, rule_json, is_enabled, updated_at
                FROM user_profile_preset
                WHERE profile_key = %s
                ORDER BY is_enabled DESC, updated_at DESC, preset_name
                """,
                (normalized_profile,),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    presets: list[dict[str, object]] = []
    for preset_name, rule_json, is_enabled, updated_at in rows:
        rule = _sanitize_profile_settings(rule_json if isinstance(rule_json, dict) else {})
        presets.append(
            {
                "preset_name": str(preset_name),
                "rule": rule,
                "is_enabled": bool(is_enabled),
                "updated_at": format_last_modified(updated_at),
            }
        )
    return presets


def upsert_profile_preset(
    settings_obj: Settings,
    profile: str,
    preset_name: str,
    rule_payload: dict[str, object],
    enabled: bool = True,
) -> None:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    normalized_name = preset_name.strip()
    if not normalized_name:
        raise ValueError("preset_name is required")
    clean_rule = _sanitize_profile_settings(rule_payload)
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO user_profile_preset (profile_key, preset_name, rule_json, is_enabled, updated_at)
                    VALUES (%s, %s, %s, %s, now())
                    ON CONFLICT (profile_key, preset_name)
                    DO UPDATE SET
                      rule_json = EXCLUDED.rule_json,
                      is_enabled = EXCLUDED.is_enabled,
                      updated_at = now()
                    """,
                    (normalized_profile, normalized_name, Json(clean_rule), bool(enabled)),
                )
    finally:
        conn.close()


def set_profile_preset_enabled(
    settings_obj: Settings,
    profile: str,
    preset_name: str,
    enabled: bool,
) -> None:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE user_profile_preset
                    SET is_enabled = %s, updated_at = now()
                    WHERE profile_key = %s
                      AND preset_name = %s
                    """,
                    (bool(enabled), normalized_profile, preset_name.strip()),
                )
    finally:
        conn.close()


def delete_profile_preset(settings_obj: Settings, profile: str, preset_name: str) -> None:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    DELETE FROM user_profile_preset
                    WHERE profile_key = %s
                      AND preset_name = %s
                    """,
                    (normalized_profile, preset_name.strip()),
                )
    finally:
        conn.close()


def rename_profile_preset(
    settings_obj: Settings,
    profile: str,
    old_name: str,
    new_name: str,
) -> None:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    old_value = old_name.strip()
    new_value = new_name.strip()
    if not old_value or not new_value:
        raise ValueError("preset name is required")
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE user_profile_preset
                    SET preset_name = %s, updated_at = now()
                    WHERE profile_key = %s
                      AND preset_name = %s
                    """,
                    (new_value, normalized_profile, old_value),
                )
    finally:
        conn.close()


def duplicate_profile_preset(
    settings_obj: Settings,
    profile: str,
    source_name: str,
    target_name: str,
) -> None:
    _ensure_profile_presets_table(settings_obj)
    normalized_profile = _normalize_user_profile(profile)
    src = source_name.strip()
    tgt = target_name.strip()
    if not src or not tgt:
        raise ValueError("preset name is required")
    conn = _connect_db(settings_obj)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO user_profile_preset (profile_key, preset_name, rule_json, is_enabled, updated_at)
                    SELECT profile_key, %s, rule_json, is_enabled, now()
                    FROM user_profile_preset
                    WHERE profile_key = %s
                      AND preset_name = %s
                    ON CONFLICT (profile_key, preset_name)
                    DO UPDATE SET
                      rule_json = EXCLUDED.rule_json,
                      is_enabled = EXCLUDED.is_enabled,
                      updated_at = now()
                    """,
                    (tgt, normalized_profile, src),
                )
    finally:
        conn.close()


def fetch_cpe_autocomplete_suggestions(
    settings_obj: Settings,
    vendor_prefix: str,
    product_prefix: str,
    version_prefix: str,
    max_items: int = 10,
) -> dict[str, list[str]]:
    vendor_key = vendor_prefix.strip().lower()
    product_key = product_prefix.strip().lower()
    version_key = version_prefix.strip().lower()
    limit = max(1, min(int(max_items), 20))
    vendors: list[str] = []
    products: list[str] = []
    versions: list[str] = []

    conn = _connect_db(settings_obj)
    try:
        with conn.cursor() as cur:
            if len(vendor_key) >= 2:
                cur.execute(
                    """
                    SELECT DISTINCT cc.vendor
                    FROM cve_cpe AS cc
                    WHERE cc.vulnerable = TRUE
                      AND cc.vendor ILIKE %s
                    ORDER BY cc.vendor
                    LIMIT %s
                    """,
                    (f"{vendor_key}%", limit),
                )
                vendors = [str(row[0]) for row in cur.fetchall() if row and row[0]]

            if len(product_key) >= 2:
                if vendor_key:
                    cur.execute(
                        """
                        SELECT DISTINCT cc.product
                        FROM cve_cpe AS cc
                        WHERE cc.vulnerable = TRUE
                          AND LOWER(cc.vendor) = %s
                          AND cc.product ILIKE %s
                        ORDER BY cc.product
                        LIMIT %s
                        """,
                        (vendor_key, f"{product_key}%", limit),
                    )
                else:
                    cur.execute(
                        """
                        SELECT DISTINCT cc.product
                        FROM cve_cpe AS cc
                        WHERE cc.vulnerable = TRUE
                          AND cc.product ILIKE %s
                        ORDER BY cc.product
                        LIMIT %s
                        """,
                        (f"{product_key}%", limit),
                    )
                products = [str(row[0]) for row in cur.fetchall() if row and row[0]]

            if len(version_key) >= 1 and vendor_key and product_key:
                cur.execute(
                    """
                    SELECT DISTINCT cc.version
                    FROM cve_cpe AS cc
                    WHERE cc.vulnerable = TRUE
                      AND LOWER(cc.vendor) = %s
                      AND LOWER(cc.product) = %s
                      AND cc.version IS NOT NULL
                      AND cc.version <> ''
                      AND cc.version ILIKE %s
                    ORDER BY cc.version
                    LIMIT %s
                    """,
                    (vendor_key, product_key, f"{version_key}%", limit),
                )
                versions = [str(row[0]) for row in cur.fetchall() if row and row[0]]
    finally:
        conn.close()

    return {
        "vendors": vendors,
        "products": products,
        "versions": versions,
    }


def fetch_cpe_preview_rows(
    settings_obj: Settings,
    vendor_value: str,
    product_value: str,
    version_value: str,
    limit: int = 12,
) -> list[dict[str, str]]:
    vendor_key = vendor_value.strip().lower()
    product_key = product_value.strip().lower()
    version_key = version_value.strip().lower()
    max_rows = max(1, min(int(limit), 20))
    where_clauses = ["cc.vulnerable = TRUE"]
    params: list[object] = []
    if vendor_key:
        where_clauses.append("cc.vendor ILIKE %s")
        params.append(f"{vendor_key}%")
    if product_key:
        where_clauses.append("cc.product ILIKE %s")
        params.append(f"{product_key}%")
    if version_key:
        where_clauses.append("COALESCE(cc.version, '') ILIKE %s")
        params.append(f"{version_key}%")

    conn = _connect_db(settings_obj)
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT DISTINCT cc.vendor, cc.product, COALESCE(NULLIF(cc.version, ''), '-') AS version
                FROM cve_cpe AS cc
                WHERE {' AND '.join(where_clauses)}
                ORDER BY cc.vendor, cc.product, version
                LIMIT %s
                """,
                [*params, max_rows],
            )
            rows = cur.fetchall()
    finally:
        conn.close()
    return [
        {"vendor": str(vendor), "product": str(product), "version": str(version)}
        for vendor, product, version in rows
    ]


def _build_count_cache_key(
    product: str,
    vendor: str,
    keyword: str,
    selected_impacts: list[str],
    min_cvss: float,
    last_modified_start_raw: str,
    last_modified_end_raw: str,
    cpe_missing_only: bool,
    selected_cpe_objects: list[str] | None = None,
) -> str:
    return "|".join(
        [
            product.lower(),
            vendor.lower(),
            keyword.lower(),
            ",".join(sorted(value.lower() for value in selected_impacts)),
            f"{min_cvss:.1f}",
            last_modified_start_raw,
            last_modified_end_raw,
            "1" if cpe_missing_only else "0",
            ",".join(sorted(value.lower() for value in (selected_cpe_objects or []))),
        ]
    )


def _get_cached_count(key: str) -> int | None:
    cached = _count_cache.get(key)
    if not cached:
        return None
    value, ts = cached
    if time.time() - ts > COUNT_CACHE_TTL_SECONDS:
        _count_cache.pop(key, None)
        return None
    return value


def _set_cached_count(key: str, value: int) -> None:
    if len(_count_cache) >= COUNT_CACHE_MAX_ENTRIES:
        oldest_key = min(_count_cache.items(), key=lambda item: item[1][1])[0]
        _count_cache.pop(oldest_key, None)
    _count_cache[key] = (value, time.time())


def parse_datetime_local(raw_value: str) -> datetime | None:
    value = raw_value.strip()
    if not value:
        return None
    return datetime.fromisoformat(value)


def _compose_datetime_arg(param_name: str) -> str:
    date_raw = (request.args.get(f"{param_name}_date") or "").strip()
    time_raw = (request.args.get(f"{param_name}_time") or "").strip()
    raw_value = (request.args.get(param_name) or "").strip()

    # If date/time split inputs are present, they take precedence over legacy raw param.
    if date_raw or time_raw:
        if not date_raw:
            return raw_value
        if not time_raw:
            return date_raw
        match = re.fullmatch(r"(\d{1,2}):(\d{2})", time_raw)
        if not match:
            return f"{date_raw}T{time_raw}"
        hour = int(match.group(1))
        minute = int(match.group(2))
        if hour > 23 or minute > 59:
            return f"{date_raw}T{time_raw}"
        return f"{date_raw}T{hour:02d}:{minute:02d}"

    return raw_value


def _split_datetime_for_inputs(raw_value: str) -> tuple[str, str]:
    value = raw_value.strip()
    if not value:
        return "", ""
    try:
        parsed = datetime.fromisoformat(value)
        return parsed.strftime("%Y-%m-%d"), parsed.strftime("%H:%M")
    except ValueError:
        if "T" in value:
            date_part, time_part = value.split("T", 1)
            return date_part, time_part[:5]
        return value[:10], ""


def format_cvss_badge(score: object) -> tuple[str, str]:
    if score is None:
        return ("None 0.0", "cvss-none")

    try:
        value = float(score)
    except (TypeError, ValueError):
        return ("None 0.0", "cvss-none")

    value = max(0.0, min(value, 10.0))
    if value == 0.0:
        return (f"None {value:.1f}", "cvss-none")
    if value <= 3.9:
        return (f"Low {value:.1f}", "cvss-low")
    if value <= 6.9:
        return (f"Medium {value:.1f}", "cvss-medium")
    if value <= 8.9:
        return (f"High {value:.1f}", "cvss-high")
    return (f"Critical {value:.1f}", "cvss-critical")


def shorten(text: str, limit: int = 130) -> str:
    clean = " ".join(text.split())
    if len(clean) <= limit:
        return clean
    return clean[: limit - 1] + "..."


def format_cpe_for_wrap(value: object) -> str:
    text = str(value)
    tokens = re.split(r"([:/._-])", text)
    chunks: list[str] = []
    for token in tokens:
        if token in {":", "/", ".", "_", "-"}:
            chunks.append(f"{escape(token)}<wbr>")
        else:
            chunks.append(escape(token))
    return "".join(chunks)


def format_last_modified(value: object) -> str:
    if isinstance(value, datetime):
        base = value.strftime("%Y-%m-%d %H:%M:%S")
        offset = value.utcoffset()
        if offset is None:
            return base
        total_minutes = int(offset.total_seconds() // 60)
        if total_minutes == 0:
            return base
        sign = "+" if total_minutes >= 0 else "-"
        abs_minutes = abs(total_minutes)
        hours = abs_minutes // 60
        minutes = abs_minutes % 60
        tz_text = f"UTC{sign}{hours:02d}:{minutes:02d}"
        return f"{base} {tz_text}"
    return str(value)


def _build_menu_html(active_page: str, user_profile: str | None = None) -> str:
    profile_query = f"?user_profile={escape(_normalize_user_profile(user_profile))}" if user_profile else ""
    return (
        "<nav class='top-menu'>"
        f"<a class='menu-link {'active' if active_page == 'search' else ''}' href='/{profile_query}'>검색</a>"
        f"<a class='menu-link {'active' if active_page == 'daily' else ''}' href='/daily{profile_query}'>일일 검토</a>"
        f"<a class='menu-link {'active' if active_page == 'settings' else ''}' href='/settings{profile_query}'>설정</a>"
        "</nav>"
    )


@app.get("/export.xlsx")
def export_xlsx() -> object:
    sort_key_param = request.args.get("sort_key")
    product = (request.args.get("product") or "").strip()
    vendor = (request.args.get("vendor") or "").strip()
    keyword = (request.args.get("keyword") or "").strip()
    last_modified_start_raw = _compose_datetime_arg("last_modified_start")
    last_modified_end_raw = _compose_datetime_arg("last_modified_end")
    cpe_missing_only = request.args.get("cpe_missing_only") == "1"
    selected_impacts = [value.strip() for value in request.args.getlist("impact_type") if value.strip()]
    selected_cpe_objects = [value.strip().lower() for value in request.args.getlist("cpe_object") if value.strip()]
    export_scope = (request.args.get("export_scope") or "page").strip().lower()
    if export_scope not in {"page", "all"}:
        export_scope = "page"

    sort_map = {
        "cvss_desc": ("cvss", "desc"),
        "cvss_asc": ("cvss", "asc"),
        "last_modified_desc": ("last_modified", "desc"),
        "last_modified_asc": ("last_modified", "asc"),
    }

    min_cvss_raw = (request.args.get("min_cvss") or "0").strip()
    limit_raw = (request.args.get("limit") or "50").strip()
    page_raw = (request.args.get("page") or "1").strip()

    try:
        min_cvss = float(min_cvss_raw)
    except ValueError:
        min_cvss = 0.0

    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 50
    limit = max(1, min(limit, 500))

    try:
        page = int(page_raw)
    except ValueError:
        page = 1
    page = max(1, page)

    sort_key = (sort_key_param or "cvss_desc").strip()
    sort_by, sort_order = sort_map.get(sort_key, ("cvss", "desc"))

    try:
        last_modified_start = parse_datetime_local(last_modified_start_raw)
        last_modified_end = parse_datetime_local(last_modified_end_raw)
    except ValueError:
        return "Invalid Last Modified datetime. Use format YYYY-MM-DDTHH:MM.", 400
    if (
        last_modified_start is not None
        and last_modified_end is not None
        and last_modified_start > last_modified_end
    ):
        return "Last Modified Start must be earlier than or equal to End.", 400

    settings = load_settings(".env")
    count_cache_key = _build_count_cache_key(
        product,
        vendor,
        keyword,
        selected_impacts,
        min_cvss,
        last_modified_start_raw,
        last_modified_end_raw,
        cpe_missing_only,
        selected_cpe_objects,
    )

    rows: list[dict[str, object]] = []
    total_count = _get_cached_count(count_cache_key)
    if export_scope == "page":
        rows, _ = fetch_cves_from_db(
            settings,
            product,
            vendor or None,
            keyword or None,
            selected_impacts or None,
            min_cvss,
            limit,
            offset=(page - 1) * limit,
            sort_by=sort_by,
            sort_order=sort_order,
            last_modified_start=last_modified_start,
            last_modified_end=last_modified_end,
            cpe_missing_only=cpe_missing_only,
            cpe_objects=selected_cpe_objects or None,
            include_total_count=False,
        )
    else:
        batch_size = 1000
        offset = 0
        while True:
            batch_rows, batch_total = fetch_cves_from_db(
                settings,
                product,
                vendor or None,
                keyword or None,
                selected_impacts or None,
                min_cvss,
                batch_size,
                offset=offset,
                sort_by=sort_by,
                sort_order=sort_order,
                last_modified_start=last_modified_start,
                last_modified_end=last_modified_end,
                cpe_missing_only=cpe_missing_only,
                cpe_objects=selected_cpe_objects or None,
                include_total_count=(total_count is None and offset == 0),
            )
            if batch_total is not None:
                total_count = batch_total
                _set_cached_count(count_cache_key, batch_total)
            if not batch_rows:
                break
            rows.extend(batch_rows)
            offset += len(batch_rows)
            if total_count is not None and offset >= total_count:
                break

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CVE Results"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    filter_summary_parts: list[str] = []
    if keyword:
        filter_summary_parts.append(f"keyword={keyword}")
    if vendor:
        filter_summary_parts.append(f"vendor={vendor}")
    if product:
        filter_summary_parts.append(f"product={product}")
    if selected_impacts:
        filter_summary_parts.append(f"impact_type={', '.join(selected_impacts)}")
    if last_modified_start_raw:
        filter_summary_parts.append(f"last_modified_start={last_modified_start_raw}")
    if last_modified_end_raw:
        filter_summary_parts.append(f"last_modified_end={last_modified_end_raw}")
    if cpe_missing_only:
        filter_summary_parts.append("cpe_missing_only=1")
    if selected_cpe_objects:
        filter_summary_parts.append(f"cpe_object={', '.join(selected_cpe_objects)}")
    filter_summary_parts.append(f"min_cvss={min_cvss}")
    filter_summary_parts.append(f"limit={limit}")
    filter_summary_parts.append(f"sort={sort_key}")
    filter_summary_parts.append(f"export_scope={export_scope}")
    filter_summary = "; ".join(filter_summary_parts)

    sheet.append(["Search Time", generated_at])
    sheet.append(["Filter Summary", filter_summary])
    sheet.append([])

    headers = [
        "CVE ID",
        "CVSS Severity",
        "CVSS Score",
        "Impact Type",
        "Last Modified",
        "Description",
        "CPE Entries",
    ]
    sheet.append(headers)
    meta_label_fill = PatternFill(fill_type="solid", start_color="E8F1EF", end_color="E8F1EF")
    header_fill = PatternFill(fill_type="solid", start_color="F2EEE4", end_color="F2EEE4")
    sheet["A1"].font = Font(bold=True, color="0F6F65")
    sheet["A2"].font = Font(bold=True, color="0F6F65")
    sheet["A1"].fill = meta_label_fill
    sheet["A2"].fill = meta_label_fill
    sheet["B1"].alignment = Alignment(horizontal="left")
    sheet["B2"].alignment = Alignment(horizontal="left", wrap_text=True)

    header_row = 4
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="2F3F45")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        severity_label, _ = format_cvss_badge(row.get("cvss_score"))
        severity = severity_label.split(" ", 1)[0]
        score_value = row.get("cvss_score")
        score_text = "0.0" if score_value is None else str(score_value)
        cpe_entries = row.get("cpe_entries") or []
        sheet.append(
            [
                str(row.get("id", "UNKNOWN")),
                severity,
                score_text,
                str(row.get("vuln_type", "Other")),
                format_last_modified(row.get("last_modified_at", "N/A")),
                str(row.get("description", "")),
                "\n".join(str(cpe) for cpe in cpe_entries) if cpe_entries else "",
            ]
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 24
    sheet.column_dimensions["F"].width = 90
    sheet.column_dimensions["G"].width = 70

    for row_idx in range(5, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=6).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_idx, column=7).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_idx, column=1).alignment = Alignment(vertical="top")
        sheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="top")
        sheet.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="top")
        sheet.cell(row=row_idx, column=4).alignment = Alignment(vertical="top")
        sheet.cell(row=row_idx, column=5).alignment = Alignment(vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    scope_text = "all" if export_scope == "all" else f"page{page}"
    filename = f"cve_export_{scope_text}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/settings", methods=["GET", "POST"])
def settings_page() -> str | object:
    user_profile = _normalize_user_profile(request.values.get("user_profile"))
    save_error = ""
    saved_notice = request.args.get("saved") == "1"
    reset_notice = request.args.get("reset") == "1"
    preview_notice = ""
    presets: list[dict[str, object]] = []

    app_settings: Settings | None = None
    profile_settings = dict(DEFAULT_PROFILE_SETTINGS)
    try:
        app_settings = load_settings(".env")
        profile_settings = fetch_profile_settings(app_settings, user_profile)
        presets = fetch_profile_presets(app_settings, user_profile)
    except Exception as exc:  # pragma: no cover
        save_error = f"설정 로딩 실패: {exc}"

    if request.method == "POST":
        action = (request.form.get("action") or "save").strip().lower()
        if action == "reset":
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 복원할 수 없습니다."
            else:
                try:
                    upsert_profile_settings(app_settings, user_profile, dict(DEFAULT_PROFILE_SETTINGS))
                    return redirect(f"/settings?user_profile={user_profile}&reset=1")
                except Exception as exc:  # pragma: no cover
                    save_error = f"기본값 복원 실패: {exc}"
        elif action == "save_preset":
            preset_name = (request.form.get("preset_name") or "").strip()
            if not preset_name:
                save_error = "프리셋 이름을 입력하세요."
            elif app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 프리셋을 저장할 수 없습니다."
            else:
                try:
                    candidate_payload = {
                        "vendor": (request.form.get("vendor") or "").strip(),
                        "product": (request.form.get("product") or "").strip(),
                        "keyword": (request.form.get("keyword") or "").strip(),
                        "cpe_objects_catalog": (request.form.get("cpe_objects_catalog") or "").strip(),
                        "min_cvss": (request.form.get("min_cvss") or "").strip(),
                        "limit": (request.form.get("limit") or "").strip(),
                        "sort_key": (request.form.get("sort_key") or "").strip(),
                        "last_modified_lookback_days": (request.form.get("last_modified_lookback_days") or "").strip(),
                        "daily_review_window_days": (request.form.get("daily_review_window_days") or "").strip(),
                        "daily_review_limit": (request.form.get("daily_review_limit") or "").strip(),
                        "cpe_missing_only": request.form.get("cpe_missing_only") == "1",
                        "impact_type": [value.strip() for value in request.form.getlist("impact_type") if value.strip()],
                    }
                    upsert_profile_preset(app_settings, user_profile, preset_name, candidate_payload, enabled=True)
                    preview_notice = f"프리셋 '{preset_name}' 저장 완료"
                except Exception as exc:  # pragma: no cover
                    save_error = f"프리셋 저장 실패: {exc}"
        elif action == "toggle_preset":
            preset_name = (request.form.get("preset_name") or "").strip()
            enabled_value = request.form.get("preset_enabled") == "1"
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 프리셋 상태를 변경할 수 없습니다."
            else:
                try:
                    set_profile_preset_enabled(app_settings, user_profile, preset_name, enabled_value)
                    preview_notice = f"프리셋 '{preset_name}' 상태 변경 완료"
                except Exception as exc:  # pragma: no cover
                    save_error = f"프리셋 상태 변경 실패: {exc}"
        elif action == "delete_preset":
            preset_name = (request.form.get("preset_name") or "").strip()
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 프리셋을 삭제할 수 없습니다."
            else:
                try:
                    delete_profile_preset(app_settings, user_profile, preset_name)
                    preview_notice = f"프리셋 '{preset_name}' 삭제 완료"
                except Exception as exc:  # pragma: no cover
                    save_error = f"프리셋 삭제 실패: {exc}"
        elif action == "rename_preset":
            preset_name = (request.form.get("preset_name") or "").strip()
            rename_to = (request.form.get("rename_to") or "").strip()
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 프리셋 이름을 변경할 수 없습니다."
            elif not rename_to:
                save_error = "변경할 프리셋 이름을 입력하세요."
            else:
                try:
                    rename_profile_preset(app_settings, user_profile, preset_name, rename_to)
                    preview_notice = f"프리셋 '{preset_name}' 이름 변경 완료"
                except Exception as exc:  # pragma: no cover
                    save_error = f"프리셋 이름 변경 실패: {exc}"
        elif action == "duplicate_preset":
            preset_name = (request.form.get("preset_name") or "").strip()
            duplicate_to = (request.form.get("duplicate_to") or "").strip()
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 프리셋을 복제할 수 없습니다."
            elif not duplicate_to:
                save_error = "복제 대상 프리셋 이름을 입력하세요."
            else:
                try:
                    duplicate_profile_preset(app_settings, user_profile, preset_name, duplicate_to)
                    preview_notice = f"프리셋 '{preset_name}' 복제 완료"
                except Exception as exc:  # pragma: no cover
                    save_error = f"프리셋 복제 실패: {exc}"

        payload = {
            "vendor": (request.form.get("vendor") or "").strip(),
            "product": (request.form.get("product") or "").strip(),
            "keyword": (request.form.get("keyword") or "").strip(),
            "cpe_objects_catalog": (request.form.get("cpe_objects_catalog") or "").strip(),
            "min_cvss": (request.form.get("min_cvss") or "").strip(),
            "limit": (request.form.get("limit") or "").strip(),
            "sort_key": (request.form.get("sort_key") or "").strip(),
            "last_modified_lookback_days": (request.form.get("last_modified_lookback_days") or "").strip(),
            "daily_review_window_days": (request.form.get("daily_review_window_days") or "").strip(),
            "daily_review_limit": (request.form.get("daily_review_limit") or "").strip(),
            "cpe_missing_only": request.form.get("cpe_missing_only") == "1",
            "impact_type": [value.strip() for value in request.form.getlist("impact_type") if value.strip()],
        }
        if action == "save":
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 저장할 수 없습니다."
                profile_settings = _sanitize_profile_settings(payload)
            else:
                try:
                    upsert_profile_settings(app_settings, user_profile, payload)
                    return redirect(f"/settings?user_profile={user_profile}&saved=1")
                except Exception as exc:  # pragma: no cover
                    save_error = f"설정 저장 실패: {exc}"
                    profile_settings = _sanitize_profile_settings(payload)
        elif action == "preview":
            profile_settings = _sanitize_profile_settings(payload)
            if app_settings is None:
                save_error = "DB 연결 정보를 불러올 수 없어 미리보기를 실행할 수 없습니다."
            else:
                try:
                    now_local = datetime.now().replace(second=0, microsecond=0)
                    window_days = int(profile_settings["daily_review_window_days"])
                    start_dt = now_local - timedelta(days=window_days)
                    _, preview_total = fetch_cves_from_db(
                        app_settings,
                        str(profile_settings["product"]) or None,
                        str(profile_settings["vendor"]) or None,
                        str(profile_settings["keyword"]) or None,
                        list(profile_settings["impact_type"]) or None,
                        float(profile_settings["min_cvss"]),
                        limit=1,
                        offset=0,
                        sort_by="last_modified",
                        sort_order="desc",
                        last_modified_start=start_dt,
                        last_modified_end=now_local,
                        cpe_missing_only=bool(profile_settings["cpe_missing_only"]),
                        cpe_objects=list(profile_settings["cpe_objects_catalog"]) or None,
                        include_total_count=True,
                    )
                    preview_notice = (
                        f"미리보기: 최근 {window_days}일 기준 예상 대상 {int(preview_total or 0)}건"
                    )
                except Exception as exc:  # pragma: no cover
                    save_error = f"미리보기 실패: {exc}"
        if app_settings is not None:
            presets = fetch_profile_presets(app_settings, user_profile)

    impact_options_html = "".join(
        "<label class='impact-option'>"
        f"<input type='checkbox' name='impact_type' value='{escape(option)}' "
        f"{'checked' if option in profile_settings['impact_type'] else ''}>"
        f"<span>{escape(option)}</span>"
        "</label>"
        for option in IMPACT_TYPE_OPTIONS
    )

    save_notice_html = "<p class='ok-msg'>저장되었습니다.</p>" if saved_notice and not save_error else ""
    reset_notice_html = "<p class='ok-msg'>기본값으로 복원되었습니다.</p>" if reset_notice and not save_error else ""
    preview_notice_html = f"<p class='ok-msg'>{escape(preview_notice)}</p>" if preview_notice and not save_error else ""
    error_html = f"<p class='error-msg'>{escape(save_error)}</p>" if save_error else ""
    menu_html = _build_menu_html("settings", user_profile=user_profile)
    cpe_catalog_json = json.dumps(profile_settings["cpe_objects_catalog"])
    preset_rows_html = "".join(
        (
            "<tr>"
            f"<td>{escape(str(item['preset_name']))}</td>"
            f"<td><span class='preset-status {'on' if item['is_enabled'] else 'off'}'>{'ON' if item['is_enabled'] else 'OFF'}</span></td>"
            f"<td>{escape(str(item['rule'].get('vendor') or '-'))}</td>"
            f"<td>{escape(str(item['rule'].get('product') or '-'))}</td>"
            f"<td>{len(item['rule'].get('cpe_objects_catalog', []))}</td>"
            f"<td>{escape(str(item.get('updated_at', '-')))}</td>"
            "<td>"
            "<form method='post' class='preset-action-inline'>"
            f"<input type='hidden' name='user_profile' value='{escape(user_profile)}'>"
            "<input type='hidden' name='action' value='toggle_preset'>"
            f"<input type='hidden' name='preset_name' value='{escape(str(item['preset_name']))}'>"
            f"<input type='hidden' name='preset_enabled' value='{'0' if item['is_enabled'] else '1'}'>"
            f"<button class='btn preset-toggle-btn {'preset-disable-btn' if item['is_enabled'] else 'preset-enable-btn'}' type='submit'>{'비활성화' if item['is_enabled'] else '활성화'}</button>"
            "</form>"
            "<form method='post' class='preset-action-inline'>"
            f"<input type='hidden' name='user_profile' value='{escape(user_profile)}'>"
            "<input type='hidden' name='action' value='rename_preset'>"
            f"<input type='hidden' name='preset_name' value='{escape(str(item['preset_name']))}'>"
            "<input class='preset-mini-input' name='rename_to' placeholder='새 이름'>"
            "<button class='btn ghost' type='submit'>이름변경</button>"
            "</form>"
            "<form method='post' class='preset-action-inline'>"
            f"<input type='hidden' name='user_profile' value='{escape(user_profile)}'>"
            "<input type='hidden' name='action' value='duplicate_preset'>"
            f"<input type='hidden' name='preset_name' value='{escape(str(item['preset_name']))}'>"
            "<input class='preset-mini-input' name='duplicate_to' placeholder='복제 이름'>"
            "<button class='btn ghost' type='submit'>복제</button>"
            "</form>"
            "<form method='post' class='preset-action-inline'>"
            f"<input type='hidden' name='user_profile' value='{escape(user_profile)}'>"
            "<input type='hidden' name='action' value='delete_preset'>"
            f"<input type='hidden' name='preset_name' value='{escape(str(item['preset_name']))}'>"
            "<button class='btn preset-delete-btn' type='submit'>삭제</button>"
            "</form>"
            "</td>"
            "</tr>"
        )
        for item in presets
    )
    if not preset_rows_html:
        preset_rows_html = "<tr><td colspan='7'>등록된 프리셋 없음</td></tr>"
    return f"""
<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CVE Settings</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg: #f7f4ee;
      --panel: #fffdf8;
      --ink: #1e2b31;
      --muted: #5e6c73;
      --line: #d7d5cc;
      --accent: #c2482e;
      --accent-2: #0f6f65;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Space Grotesk", "Pretendard", "Noto Sans KR", sans-serif;
      color: var(--ink);
      background: radial-gradient(circle at 0% 0%, #fff9ef 0, #f7f4ee 58%);
    }}
    .wrap {{ width: min(1600px, 90vw); margin: 34px auto 54px; }}
    .top-menu {{
      display: flex;
      gap: 10px;
      margin-bottom: 14px;
    }}
    .menu-link {{
      text-decoration: none;
      color: var(--ink);
      border: 1px solid var(--line);
      background: #fff8ed;
      border-radius: 999px;
      padding: 8px 14px;
      font-size: 13px;
      font-weight: 700;
    }}
    .menu-link.active {{
      color: #fff;
      background: var(--accent-2);
      border-color: var(--accent-2);
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 16px;
      padding: 20px;
    }}
    .header-row {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      margin-bottom: 12px;
    }}
    .profile-tabs {{
      display: flex;
      gap: 8px;
    }}
    .tab {{
      text-decoration: none;
      border: 1px solid var(--line);
      color: var(--ink);
      padding: 6px 12px;
      border-radius: 999px;
      font-size: 13px;
      font-weight: 700;
      background: #fff;
    }}
    .tab.active {{
      color: #fff;
      background: var(--accent);
      border-color: var(--accent);
    }}
    form {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px 16px;
    }}
    label {{
      display: block;
      font-size: 12px;
      font-weight: 700;
      color: var(--muted);
      margin-bottom: 6px;
    }}
    input, select {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px 12px;
      font: inherit;
      background: #fff;
    }}
    .full {{ grid-column: 1 / -1; }}
    .impact-box {{
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px 12px;
      background: #fff;
      max-height: 180px;
      overflow: auto;
      display: grid;
      gap: 8px;
    }}
    .impact-option {{
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 13px;
      color: var(--ink);
    }}
    .impact-option input {{
      width: auto;
      margin: 0;
    }}
    .checkbox-row {{
      display: flex;
      align-items: center;
      gap: 8px;
      min-height: 42px;
    }}
    .checkbox-row label {{
      margin: 0;
      color: var(--ink);
      font-size: 13px;
      font-weight: 600;
    }}
    .actions {{
      grid-column: 1 / -1;
      display: flex;
      justify-content: flex-end;
      align-items: center;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 6px;
    }}
    .preset-name-input {{
      width: 180px;
      min-width: 160px;
    }}
    .preset-mini-input {{
      width: 120px;
      min-width: 110px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px 10px;
      font: inherit;
      background: #fff;
    }}
    .preset-action-inline {{
      display: inline-flex;
      gap: 6px;
      align-items: center;
      margin-right: 6px;
      margin-bottom: 6px;
    }}
    .preset-status {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 46px;
      border-radius: 999px;
      border: 1px solid transparent;
      font-size: 11px;
      font-weight: 700;
      padding: 2px 8px;
    }}
    .preset-status.on {{
      background: #e7f6ef;
      color: #0e6a4c;
      border-color: #9ecfb9;
    }}
    .preset-status.off {{
      background: #fff0ea;
      color: #8d3a21;
      border-color: #efb8a3;
    }}
    .btn {{
      text-decoration: none;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px 14px;
      font: inherit;
      font-size: 13px;
      font-weight: 700;
      cursor: pointer;
      background: #fff;
      color: var(--ink);
    }}
    .btn.ghost {{
      background: #f7f8f8;
      border-color: #d8dedb;
      color: #244149;
    }}
    .btn.warn {{
      background: #fff2eb;
      border-color: #efc0a8;
      color: #8a3d1e;
    }}
    .preset-enable-btn {{
      background: #e7f6ef;
      border-color: #9ecfb9;
      color: #0e6a4c;
    }}
    .preset-toggle-btn {{
      min-width: 88px;
      text-align: center;
    }}
    .preset-disable-btn {{
      background: #fff0ea;
      border-color: #efb8a3;
      color: #8d3a21;
    }}
    .preset-delete-btn {{
      background: #fff6f2;
      border-color: #e8c8b9;
      color: #8b4432;
    }}
    .btn.primary {{
      background: var(--accent-2);
      color: #fff;
      border-color: var(--accent-2);
    }}
    .ok-msg {{
      margin: 0 0 10px;
      color: #0c6d57;
      font-size: 13px;
      font-weight: 700;
    }}
    .error-msg {{
      margin: 0 0 10px;
      color: #ad3427;
      font-size: 13px;
      font-weight: 700;
    }}
    @media (max-width: 820px) {{
      form {{ grid-template-columns: 1fr; }}
      .header-row {{ flex-direction: column; align-items: flex-start; }}
      .actions {{ justify-content: stretch; }}
      .btn {{ flex: 1; text-align: center; min-width: 140px; }}
      .preset-name-input {{ width: 100%; min-width: 0; }}
    }}
  </style>
</head>
<body>
  <main class="wrap">
    {menu_html}
    <section class="panel">
      <div class="header-row">
        <div>
          <h1 style="margin:0 0 4px;font-size:24px;">사용자 설정</h1>
          <p style="margin:0;color:var(--muted);font-size:13px;">프로필별 기본 검색값을 저장합니다.</p>
        </div>
        <div class="profile-tabs">
          <a class="tab {'active' if user_profile == 'hq' else ''}" href="/settings?user_profile=hq">본사</a>
          <a class="tab {'active' if user_profile == 'jaehwa' else ''}" href="/settings?user_profile=jaehwa">재화</a>
        </div>
      </div>
      {save_notice_html}
      {reset_notice_html}
      {preview_notice_html}
      {error_html}
      <form method="post">
        <input type="hidden" name="user_profile" value="{escape(user_profile)}">
        <div>
          <label for="vendor">기본 Vendor</label>
          <input id="vendor" name="vendor" value="{escape(str(profile_settings['vendor']))}" placeholder="e.g. ivanti">
        </div>
        <div>
          <label for="product">기본 Product (쉼표로 OR)</label>
          <input id="product" name="product" value="{escape(str(profile_settings['product']))}" placeholder="e.g. endpoint_manager_mobile, pulse_connect_secure">
        </div>
        <div>
          <label for="keyword">기본 Keyword (쉼표로 OR)</label>
          <input id="keyword" name="keyword" value="{escape(str(profile_settings['keyword']))}" placeholder="e.g. ssl, auth bypass, rce">
        </div>
        <div class="full">
          <label>CPE 객체 목록 (vendor:product[:version])</label>
          <input type="hidden" id="cpe_objects_catalog" name="cpe_objects_catalog" value="{escape(chr(10).join(profile_settings['cpe_objects_catalog']))}">
          <div class="impact-box" id="cpe-catalog-list" style="max-height:none;min-height:60px;"></div>
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:8px;margin-top:8px;">
            <input id="cpe_vendor" list="cpe_vendor_suggestions" placeholder="vendor (e.g. ivanti)">
            <input id="cpe_product" list="cpe_product_suggestions" placeholder="product (e.g. pulse_connect_secure)">
            <input id="cpe_version" list="cpe_version_suggestions" placeholder="version (optional)">
            <button class="btn" type="button" id="add-cpe-btn">추가</button>
          </div>
          <p id="cpe-input-hint" style="margin:6px 0 0;font-size:12px;color:#5e6c73;">입력 형식: vendor + product는 필수, version은 선택입니다.</p>
          <div id="cpe-preview-grid" style="margin-top:8px;border:1px solid var(--line);border-radius:10px;overflow:hidden;background:#fff;">
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr;background:#f6f8f7;font-size:12px;font-weight:700;color:#4c5a60;padding:8px 10px;">
              <span>Vendor</span><span>Product</span><span>Version</span>
            </div>
            <div id="cpe-preview-rows" style="font-size:12px;color:#2f3f45;padding:8px 10px;">추천 결과 없음</div>
          </div>
          <datalist id="cpe_vendor_suggestions"></datalist>
          <datalist id="cpe_product_suggestions"></datalist>
          <datalist id="cpe_version_suggestions"></datalist>
        </div>
        <div>
          <label for="min_cvss">기본 Min CVSS</label>
          <input id="min_cvss" name="min_cvss" type="number" min="0" max="10" step="0.1" value="{escape(str(profile_settings['min_cvss']))}">
        </div>
        <div>
          <label for="limit">기본 Limit (1-500)</label>
          <input id="limit" name="limit" type="number" min="1" max="500" step="1" value="{escape(str(profile_settings['limit']))}">
        </div>
        <div>
          <label for="sort_key">기본 정렬</label>
          <select id="sort_key" name="sort_key">
            <option value="cvss_desc" {'selected' if profile_settings['sort_key'] == 'cvss_desc' else ''}>CVSS 내림차순</option>
            <option value="cvss_asc" {'selected' if profile_settings['sort_key'] == 'cvss_asc' else ''}>CVSS 오름차순</option>
            <option value="last_modified_desc" {'selected' if profile_settings['sort_key'] == 'last_modified_desc' else ''}>Last Modified 내림차순</option>
            <option value="last_modified_asc" {'selected' if profile_settings['sort_key'] == 'last_modified_asc' else ''}>Last Modified 오름차순</option>
          </select>
        </div>
        <div>
          <label for="last_modified_lookback_days">기본 Last Modified 범위(일)</label>
          <input id="last_modified_lookback_days" name="last_modified_lookback_days" type="number" min="1" max="365" step="1" value="{escape(str(profile_settings['last_modified_lookback_days']))}">
        </div>
        <div>
          <label for="daily_review_window_days">일일 검토 기본 기간(일)</label>
          <input id="daily_review_window_days" name="daily_review_window_days" type="number" min="1" max="30" step="1" value="{escape(str(profile_settings['daily_review_window_days']))}">
        </div>
        <div>
          <label for="daily_review_limit">일일 검토 최대 건수</label>
          <input id="daily_review_limit" name="daily_review_limit" type="number" min="1" max="1000" step="1" value="{escape(str(profile_settings['daily_review_limit']))}">
        </div>
        <div class="checkbox-row">
          <input id="cpe_missing_only" name="cpe_missing_only" type="checkbox" value="1" {'checked' if profile_settings['cpe_missing_only'] else ''}>
          <label for="cpe_missing_only">기본 CPE missing only 사용</label>
        </div>
        <div class="full">
          <label>기본 Impact Type</label>
          <div class="impact-box">
            {impact_options_html}
          </div>
        </div>
        <div class="actions">
          <a class="btn ghost" href="/?user_profile={escape(user_profile)}">검색으로 이동</a>
          <button class="btn warn" type="submit" name="action" value="reset">기본값 복원</button>
          <input class="preset-name-input" name="preset_name" placeholder="프리셋 이름">
          <button class="btn ghost" type="submit" name="action" value="save_preset">현재값을 프리셋 저장</button>
          <button class="btn ghost" type="submit" name="action" value="preview">미리보기</button>
          <button class="btn primary" type="submit">저장</button>
        </div>
      </form>
      <div class="full" style="margin-top:12px;">
        <h3 style="margin:8px 0;font-size:16px;">프리셋 목록 (일일검토 OR 합집합 대상)</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
          <thead>
            <tr>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">프리셋</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">상태</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">Vendor</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">Product</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">CPE 수</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">수정시각</th>
              <th style="text-align:left;border-top:1px solid var(--line);padding:8px;">관리</th>
            </tr>
          </thead>
          <tbody>
            {preset_rows_html}
          </tbody>
        </table>
      </div>
    </section>
  </main>
</body>
<script>
  (() => {{
    const hiddenInput = document.querySelector("#cpe_objects_catalog");
    const listWrap = document.querySelector("#cpe-catalog-list");
    const addBtn = document.querySelector("#add-cpe-btn");
    const vendorInput = document.querySelector("#cpe_vendor");
    const productInput = document.querySelector("#cpe_product");
    const versionInput = document.querySelector("#cpe_version");
    const vendorList = document.querySelector("#cpe_vendor_suggestions");
    const productList = document.querySelector("#cpe_product_suggestions");
    const versionList = document.querySelector("#cpe_version_suggestions");
    const inputHint = document.querySelector("#cpe-input-hint");
    const previewRows = document.querySelector("#cpe-preview-rows");
    let catalog = {cpe_catalog_json};
    let suggestTimer = null;
    let suggestAbortController = null;
    let previewTimer = null;
    let previewAbortController = null;

    const normalize = (value) => (value || "").trim().toLowerCase();
    const setDataList = (target, items) => {{
      if (!target) return;
      target.innerHTML = (items || []).map((item) => `<option value="${{item}}"></option>`).join("");
    }};
    const fetchSuggest = async () => {{
      const vendor = normalize(vendorInput?.value);
      const product = normalize(productInput?.value);
      const version = normalize(versionInput?.value);
      const canFetch = (vendor.length >= 2) || (product.length >= 2) || (version.length >= 1 && vendor && product);
      if (!canFetch) {{
        setDataList(vendorList, []);
        setDataList(productList, []);
        setDataList(versionList, []);
        return;
      }}
      if (suggestAbortController) {{
        suggestAbortController.abort();
      }}
      suggestAbortController = new AbortController();
      const params = new URLSearchParams();
      if (vendor) params.set("vendor", vendor);
      if (product) params.set("product", product);
      if (version) params.set("version", version);
      params.set("limit", "10");
      try {{
        const response = await fetch("/api/cpe/suggest?" + params.toString(), {{
          method: "GET",
          signal: suggestAbortController.signal,
          headers: {{ "Accept": "application/json" }},
        }});
        if (!response.ok) {{
          return;
        }}
        const data = await response.json();
        setDataList(vendorList, data?.vendors || []);
        setDataList(productList, data?.products || []);
        setDataList(versionList, data?.versions || []);
      }} catch (_) {{
        // Ignore aborted or transient suggestion errors.
      }}
    }};
    const queueSuggest = () => {{
      if (suggestTimer) {{
        clearTimeout(suggestTimer);
      }}
      suggestTimer = setTimeout(fetchSuggest, 220);
    }};
    const renderPreviewRows = (rows) => {{
      if (!previewRows) return;
      if (!rows || !rows.length) {{
        previewRows.textContent = "추천 결과 없음";
        return;
      }}
      previewRows.innerHTML = rows.map((row) => {{
        const vendor = row?.vendor || "-";
        const product = row?.product || "-";
        const version = row?.version || "-";
        return `<div style="display:grid;grid-template-columns:1fr 1fr 1fr;padding:4px 0;border-top:1px dashed #e2e6e4;"><span>${{vendor}}</span><span>${{product}}</span><span>${{version}}</span></div>`;
      }}).join("");
    }};
    const fetchPreview = async () => {{
      const vendor = normalize(vendorInput?.value);
      const product = normalize(productInput?.value);
      const version = normalize(versionInput?.value);
      if (previewAbortController) {{
        previewAbortController.abort();
      }}
      previewAbortController = new AbortController();
      const params = new URLSearchParams();
      if (vendor) params.set("vendor", vendor);
      if (product) params.set("product", product);
      if (version) params.set("version", version);
      params.set("limit", "10");
      try {{
        const response = await fetch("/api/cpe/preview?" + params.toString(), {{
          method: "GET",
          signal: previewAbortController.signal,
          headers: {{ "Accept": "application/json" }},
        }});
        if (!response.ok) {{
          return;
        }}
        const data = await response.json();
        renderPreviewRows(data?.rows || []);
      }} catch (_) {{
        // Ignore aborted or transient preview errors.
      }}
    }};
    const queuePreview = () => {{
      if (previewTimer) {{
        clearTimeout(previewTimer);
      }}
      previewTimer = setTimeout(fetchPreview, 260);
    }};
    const rebuildHidden = () => {{
      if (hiddenInput) {{
        hiddenInput.value = catalog.join("\\n");
      }}
    }};
    const render = () => {{
      if (!listWrap) return;
      if (!catalog.length) {{
        listWrap.innerHTML = "<span class='impact-chip muted-chip'>등록된 CPE 객체가 없습니다.</span>";
        rebuildHidden();
        return;
      }}
      listWrap.innerHTML = catalog.map((item, idx) => {{
        return `<span class="impact-chip">${{item}} <button type="button" data-cpe-idx="${{idx}}" style="margin-left:6px;border:0;background:transparent;cursor:pointer;color:#8a2f23;">x</button></span>`;
      }}).join("");
      listWrap.querySelectorAll("[data-cpe-idx]").forEach((btn) => {{
        btn.addEventListener("click", () => {{
          const idx = Number(btn.getAttribute("data-cpe-idx") || "-1");
          if (idx < 0) return;
          catalog = catalog.filter((_, i) => i !== idx);
          render();
        }});
      }});
      rebuildHidden();
    }};

    addBtn?.addEventListener("click", () => {{
      const vendor = normalize(vendorInput?.value);
      const product = normalize(productInput?.value);
      const version = normalize(versionInput?.value);
      if (!vendor || !product) {{
        if (inputHint) {{
          inputHint.textContent = "vendor와 product를 모두 입력해야 추가됩니다.";
          inputHint.style.color = "#ad3427";
        }}
        return;
      }}
      if (inputHint) {{
        inputHint.textContent = "입력 형식: vendor + product는 필수, version은 선택입니다.";
        inputHint.style.color = "#5e6c73";
      }}
      const cpe = version ? `${{vendor}}:${{product}}:${{version}}` : `${{vendor}}:${{product}}`;
      if (!catalog.includes(cpe)) {{
        catalog.push(cpe);
      }}
      if (vendorInput) vendorInput.value = "";
      if (productInput) productInput.value = "";
      if (versionInput) versionInput.value = "";
      render();
    }});
    vendorInput?.addEventListener("input", queueSuggest);
    productInput?.addEventListener("input", queueSuggest);
    versionInput?.addEventListener("input", queueSuggest);
    vendorInput?.addEventListener("input", queuePreview);
    productInput?.addEventListener("input", queuePreview);
    versionInput?.addEventListener("input", queuePreview);
    vendorInput?.addEventListener("focus", queueSuggest);
    productInput?.addEventListener("focus", queueSuggest);
    versionInput?.addEventListener("focus", queueSuggest);
    vendorInput?.addEventListener("focus", queuePreview);
    productInput?.addEventListener("focus", queuePreview);
    versionInput?.addEventListener("focus", queuePreview);

    render();
    fetchPreview();
  }})();
</script>
</html>
"""


@app.get("/api/cpe/suggest")
def api_cpe_suggest() -> object:
    vendor = (request.args.get("vendor") or "").strip()
    product = (request.args.get("product") or "").strip()
    version = (request.args.get("version") or "").strip()
    limit_raw = (request.args.get("limit") or "10").strip()
    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 10
    limit = max(1, min(limit, 20))
    try:
        settings_obj = load_settings(".env")
        data = fetch_cpe_autocomplete_suggestions(settings_obj, vendor, product, version, max_items=limit)
        return jsonify(data)
    except Exception as exc:  # pragma: no cover
        return jsonify({"vendors": [], "products": [], "versions": [], "error": str(exc)}), 500


@app.get("/api/cpe/preview")
def api_cpe_preview() -> object:
    vendor = (request.args.get("vendor") or "").strip()
    product = (request.args.get("product") or "").strip()
    version = (request.args.get("version") or "").strip()
    limit_raw = (request.args.get("limit") or "10").strip()
    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 10
    limit = max(1, min(limit, 20))
    try:
        settings_obj = load_settings(".env")
        rows = fetch_cpe_preview_rows(settings_obj, vendor, product, version, limit=limit)
        return jsonify({"rows": rows})
    except Exception as exc:  # pragma: no cover
        return jsonify({"rows": [], "error": str(exc)}), 500


@app.get("/")
def index() -> str:
    user_profile = _normalize_user_profile(request.args.get("user_profile"))
    app_settings: Settings | None = None
    profile_defaults = dict(DEFAULT_PROFILE_SETTINGS)
    bootstrap_error = ""
    try:
        app_settings = load_settings(".env")
        profile_defaults = fetch_profile_settings(app_settings, user_profile)
    except Exception as exc:  # pragma: no cover
        bootstrap_error = f"설정 로딩 실패: {exc}"

    product = (
        (request.args.get("product") if "product" in request.args else str(profile_defaults["product"]) or "").strip()
    )
    vendor = (
        (request.args.get("vendor") if "vendor" in request.args else str(profile_defaults["vendor"]) or "").strip()
    )
    keyword = (
        (request.args.get("keyword") if "keyword" in request.args else str(profile_defaults["keyword"]) or "").strip()
    )

    user_supplied_last_modified = any(
        name in request.args
        for name in {
            "last_modified_present",
            "last_modified_start",
            "last_modified_end",
            "last_modified_start_date",
            "last_modified_start_time",
            "last_modified_end_date",
            "last_modified_end_time",
        }
    )
    if user_supplied_last_modified:
        last_modified_start_raw = _compose_datetime_arg("last_modified_start")
        last_modified_end_raw = _compose_datetime_arg("last_modified_end")
    else:
        lookback_days = int(profile_defaults["last_modified_lookback_days"])
        now_local = datetime.now().replace(second=0, microsecond=0)
        last_modified_end_raw = now_local.isoformat(timespec="minutes")
        last_modified_start_raw = (now_local - timedelta(days=lookback_days)).isoformat(timespec="minutes")
    last_modified_start_date_raw, last_modified_start_time_raw = _split_datetime_for_inputs(last_modified_start_raw)
    last_modified_end_date_raw, last_modified_end_time_raw = _split_datetime_for_inputs(last_modified_end_raw)

    if "cpe_missing_only_present" in request.args or "cpe_missing_only" in request.args:
        cpe_missing_only = request.args.get("cpe_missing_only") == "1"
    else:
        cpe_missing_only = bool(profile_defaults["cpe_missing_only"])

    if "impact_type_present" in request.args or "impact_type" in request.args:
        selected_impacts = [value.strip() for value in request.args.getlist("impact_type") if value.strip()]
    else:
        selected_impacts = list(profile_defaults["impact_type"])
    cpe_objects_catalog = [str(value).strip().lower() for value in profile_defaults.get("cpe_objects_catalog", []) if str(value).strip()]
    if "cpe_object_present" in request.args or "cpe_object" in request.args:
        selected_cpe_objects = [value.strip().lower() for value in request.args.getlist("cpe_object") if value.strip()]
    else:
        selected_cpe_objects = []
    selected_cpe_objects = [value for value in selected_cpe_objects if value in cpe_objects_catalog]

    sort_map = {
        "cvss_desc": ("cvss", "desc"),
        "cvss_asc": ("cvss", "asc"),
        "last_modified_desc": ("last_modified", "desc"),
        "last_modified_asc": ("last_modified", "asc"),
    }

    min_cvss_raw = (
        request.args.get("min_cvss")
        if "min_cvss" in request.args
        else str(profile_defaults["min_cvss"])
    )
    limit_raw = (
        request.args.get("limit")
        if "limit" in request.args
        else str(profile_defaults["limit"])
    )
    min_cvss_raw = (min_cvss_raw or "0").strip()
    limit_raw = (limit_raw or "50").strip()

    try:
        min_cvss = float(min_cvss_raw)
    except ValueError:
        min_cvss = 0.0

    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 50
    limit = max(1, min(limit, 500))

    page_raw = (request.args.get("page") or "1").strip()
    try:
        page = int(page_raw)
    except ValueError:
        page = 1
    page = max(1, page)
    offset = (page - 1) * limit

    sort_key_param = (
        request.args.get("sort_key")
        if "sort_key" in request.args
        else str(profile_defaults["sort_key"])
    )
    sort_key = (sort_key_param or "cvss_desc").strip().lower()
    sort_by, sort_order = sort_map.get(sort_key, ("cvss", "desc"))

    last_modified_start: datetime | None = None
    last_modified_end: datetime | None = None
    rows: list[dict[str, object]] = []
    total_count = 0
    error_text = ""
    checkpoint_text = "기록 없음"
    count_cache_key = _build_count_cache_key(
        product,
        vendor,
        keyword,
        selected_impacts,
        min_cvss,
        last_modified_start_raw,
        last_modified_end_raw,
        cpe_missing_only,
        selected_cpe_objects,
    )
    cached_total = _get_cached_count(count_cache_key)
    should_fetch_total_count = (page == 1) or (cached_total is None)

    try:
        if last_modified_start_raw:
            last_modified_start = datetime.fromisoformat(last_modified_start_raw)
        if last_modified_end_raw:
            last_modified_end = datetime.fromisoformat(last_modified_end_raw)
    except ValueError:
        error_text = "Invalid Last Modified datetime. Use format YYYY-MM-DDTHH:MM."

    if (
        not error_text
        and last_modified_start is not None
        and last_modified_end is not None
        and last_modified_start > last_modified_end
    ):
        error_text = "Last Modified Start must be earlier than or equal to End."

    if not error_text and not bootstrap_error:
        try:
            try:
                checkpoint_value = fetch_incremental_checkpoint(app_settings)
                checkpoint_text = format_last_modified(checkpoint_value) if checkpoint_value else "기록 없음"
            except Exception:
                checkpoint_text = "조회 실패"
            rows, total_count = fetch_cves_from_db(
                app_settings,
                product,
                vendor or None,
                keyword or None,
                selected_impacts or None,
                min_cvss,
                limit,
                offset=offset,
                sort_by=sort_by,
                sort_order=sort_order,
                last_modified_start=last_modified_start,
                last_modified_end=last_modified_end,
                cpe_missing_only=cpe_missing_only,
                cpe_objects=selected_cpe_objects or None,
                include_total_count=should_fetch_total_count,
            )
            if total_count is None:
                total_count = cached_total or 0
            else:
                _set_cached_count(count_cache_key, total_count)
        except Exception as exc:  # pragma: no cover
            error_text = str(exc)
    if bootstrap_error and not error_text:
        error_text = bootstrap_error

    row_chunks: list[str] = []
    for row in rows:
        cve_id = escape(str(row.get("id", "UNKNOWN")))
        score_label, score_class = format_cvss_badge(row.get("cvss_score"))
        score_text = escape(score_label)
        vuln_type = escape(str(row.get("vuln_type", "Other")))
        last_modified = escape(format_last_modified(row.get("last_modified_at", "N/A")))
        description = str(row.get("description", ""))
        summary = escape(shorten(description))
        full_description = escape(description)
        cpe_entries = row.get("cpe_entries") or []
        cpe_badges = "".join(
            f"<span class='cpe-chip'>{format_cpe_for_wrap(cpe_value)}</span>" for cpe_value in cpe_entries[:10]
        )
        if not cpe_badges:
            cpe_badges = "<span class='cpe-chip'>-</span>"
        cpe_for_copy = escape(", ".join(str(cpe_value) for cpe_value in cpe_entries)) if cpe_entries else "-"

        row_chunks.append(
            "<tr>"
            f"<td class='id'>{cve_id}</td>"
            f"<td class='score'><span class='cvss-chip {escape(score_class)}'>{score_text}</span></td>"
            f"<td class='lastmod'>{last_modified}</td>"
            f"<td class='vtype'>{vuln_type}</td>"
            f"<td class='desc'>{summary}</td>"
            f"<td class='cpe'><div class='cpe-wrap'>{cpe_badges}</div></td>"
            "<td class='actions'>"
            f"<button type='button' class='copy-btn' data-copy='{cve_id}'>Copy CVE</button>"
            f"<button type='button' class='copy-btn alt' data-copy='{cpe_for_copy}'>Copy CPE</button>"
            f"<button type='button' class='copy-btn view-btn' data-cve='{cve_id}' data-desc='{full_description}'>View</button>"
            "</td>"
            "</tr>"
        )
    rows_html = "".join(row_chunks)
    if not rows_html:
        rows_html = "<tr><td colspan='7'>No results</td></tr>"

    base_query: dict[str, object] = {
        "user_profile": user_profile,
        "vendor": vendor,
        "product": product,
        "keyword": keyword,
        "min_cvss": str(min_cvss),
        "limit": str(limit),
        "page": str(page),
        "sort_key": sort_key,
    }
    if "impact_type_present" in request.args:
        base_query["impact_type_present"] = "1"
    if "cpe_missing_only_present" in request.args:
        base_query["cpe_missing_only_present"] = "1"
    if "last_modified_present" in request.args:
        base_query["last_modified_present"] = "1"
    if "cpe_object_present" in request.args:
        base_query["cpe_object_present"] = "1"
    if last_modified_start_raw:
        base_query["last_modified_start"] = last_modified_start_raw
    if last_modified_end_raw:
        base_query["last_modified_end"] = last_modified_end_raw
    if selected_impacts:
        base_query["impact_type"] = selected_impacts
    if cpe_missing_only:
        base_query["cpe_missing_only"] = "1"
    if selected_cpe_objects:
        base_query["cpe_object"] = selected_cpe_objects

    def build_sort_href(target: str) -> str:
        if target == "cvss":
            next_key = "cvss_asc" if sort_key == "cvss_desc" else "cvss_desc"
        else:
            next_key = "last_modified_asc" if sort_key == "last_modified_desc" else "last_modified_desc"
        query = dict(base_query)
        query["sort_key"] = next_key
        return "?" + urlencode(query, doseq=True)

    cvss_sort_href = build_sort_href("cvss")
    last_modified_sort_href = build_sort_href("last_modified")
    cvss_sort_marker = "▼" if sort_key == "cvss_desc" else ("▲" if sort_key == "cvss_asc" else "")
    last_modified_sort_marker = (
        "▼" if sort_key == "last_modified_desc" else ("▲" if sort_key == "last_modified_asc" else "")
    )

    impact_options_html = "".join(
        "<label class='impact-option'>"
        f"<input type='checkbox' name='impact_type' value='{escape(option)}' {'checked' if option in selected_impacts else ''}>"
        f"<span>{escape(option)}</span>"
        "</label>"
        for option in IMPACT_TYPE_OPTIONS
    )
    impact_summary = "All impact types" if not selected_impacts else f"Impact Type ({len(selected_impacts)} selected)"
    impact_selected_html = (
        "".join(f"<span class='impact-chip'>{escape(value)}</span>" for value in selected_impacts)
        if selected_impacts
        else "<span class='impact-chip muted-chip'>No filter</span>"
    )
    cpe_object_options_html = "".join(
        "<label class='impact-option'>"
        f"<input type='checkbox' name='cpe_object' value='{escape(cpe_obj)}' {'checked' if cpe_obj in selected_cpe_objects else ''}>"
        f"<span>{escape(cpe_obj)}</span>"
        "</label>"
        for cpe_obj in cpe_objects_catalog
    )
    cpe_object_summary = (
        "No CPE object configured"
        if not cpe_objects_catalog
        else (
            "All configured CPE objects" if not selected_cpe_objects else f"CPE Object ({len(selected_cpe_objects)} selected)"
        )
    )
    cpe_object_selected_html = (
        "".join(f"<span class='impact-chip'>{escape(value)}</span>" for value in selected_cpe_objects)
        if selected_cpe_objects
        else "<span class='impact-chip muted-chip'>No filter</span>"
    )
    cpe_missing_checked = "checked" if cpe_missing_only else ""
    total_pages = max(1, math.ceil(total_count / limit)) if total_count > 0 else 1
    if page > total_pages:
        page = total_pages
    page_start = max(1, page - 2)
    page_end = min(total_pages, page + 2)
    pager_links: list[str] = []
    for page_no in range(page_start, page_end + 1):
        page_query = dict(base_query)
        page_query["page"] = str(page_no)
        page_href = "?" + urlencode(page_query, doseq=True)
        if page_no == page:
            pager_links.append(f"<span class='page-link current'>{page_no}</span>")
        else:
            pager_links.append(f"<a class='page-link' href='{escape(page_href)}'>{page_no}</a>")
    prev_href = ""
    next_href = ""
    if page > 1:
        prev_query = dict(base_query)
        prev_query["page"] = str(page - 1)
        prev_href = "?" + urlencode(prev_query, doseq=True)
    if page < total_pages:
        next_query = dict(base_query)
        next_query["page"] = str(page + 1)
        next_href = "?" + urlencode(next_query, doseq=True)
    prev_link_html = (
        f"<a class='page-link' href='{escape(prev_href)}'>Prev</a>"
        if prev_href
        else "<span class='page-link disabled'>Prev</span>"
    )
    next_link_html = (
        f"<a class='page-link' href='{escape(next_href)}'>Next</a>"
        if next_href
        else "<span class='page-link disabled'>Next</span>"
    )
    pager_html = (
        "<div class='pager'>"
        f"{prev_link_html}"
        f"{''.join(pager_links)}"
        f"{next_link_html}"
        "</div>"
    )

    error_html = f"<p class='error'>Error: {escape(error_text)}</p>" if error_text else ""
    menu_html = _build_menu_html("search", user_profile=user_profile)

    return f"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CVE Query</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg: #f7f4ee;
      --panel: #fffdf8;
      --ink: #1e2b31;
      --muted: #5e6c73;
      --line: #d7d5cc;
      --accent: #c2482e;
      --accent-2: #0f6f65;
      --shadow: 0 14px 36px rgba(30, 43, 49, 0.12);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: 'Space Grotesk', 'Segoe UI', sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 8% 0%, rgba(194, 72, 46, 0.16), transparent 36%),
        radial-gradient(circle at 100% 100%, rgba(15, 111, 101, 0.2), transparent 45%),
        var(--bg);
      min-height: 100vh;
    }}
    .wrap {{
      width: min(1600px, 90vw);
      margin: 34px auto 54px;
      animation: rise 280ms ease-out;
    }}
    .top-menu {{
      display: flex;
      gap: 10px;
      margin-bottom: 12px;
    }}
    .menu-link {{
      text-decoration: none;
      color: var(--ink);
      border: 1px solid var(--line);
      background: #fff8ed;
      border-radius: 999px;
      padding: 8px 14px;
      font-size: 13px;
      font-weight: 700;
    }}
    .menu-link.active {{
      color: #fff;
      background: var(--accent-2);
      border-color: var(--accent-2);
    }}
    .hero {{
      margin-bottom: 14px;
      padding: 20px 22px;
      border: 1px solid var(--line);
      border-radius: 18px;
      background: linear-gradient(135deg, #fffdf8, #fef8ed);
      box-shadow: var(--shadow);
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 14px;
      flex-wrap: wrap;
    }}
    .hero-copy {{ min-width: 260px; }}
    h1 {{
      margin: 0;
      font-size: clamp(24px, 3vw, 34px);
      letter-spacing: 0.2px;
    }}
    .sub {{
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 14px;
    }}
    .checkpoint-badge {{
      margin-left: auto;
      border: 1px solid #c8d8d3;
      border-radius: 11px;
      padding: 8px 10px;
      background: rgba(255, 255, 255, 0.8);
      font-size: 12px;
      color: #27424a;
      line-height: 1.35;
      text-align: right;
    }}
    .checkpoint-badge strong {{
      display: block;
      color: #123a44;
      font-size: 11px;
      letter-spacing: 0.2px;
      text-transform: uppercase;
      margin-bottom: 2px;
    }}
    .panel {{
      border: 1px solid var(--line);
      border-radius: 18px;
      background: var(--panel);
      box-shadow: var(--shadow);
      overflow: visible;
    }}
    .panel > form {{
      padding: 16px;
      display: grid;
      grid-template-columns:
        minmax(150px, 1.25fr)
        minmax(180px, 1.5fr)
        minmax(90px, 0.65fr)
        minmax(220px, 1.9fr)
        minmax(90px, 0.65fr)
        minmax(120px, 0.85fr);
      gap: 10px;
      align-items: start;
      background: linear-gradient(180deg, #fffdf8, #fff9ef);
      border-bottom: 1px solid var(--line);
    }}
    .field-lastmod-start {{ grid-column: 1 / 2; }}
    .field-lastmod-end {{ grid-column: 2 / 3; }}
    .datetime-parts {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) 88px;
      gap: 6px;
    }}
    .datetime-time {{
      text-align: center;
      font-variant-numeric: tabular-nums;
    }}
    .field-keyword {{ grid-column: 1 / 3; }}
    .field-vendor {{ grid-column: 3 / 5; }}
    .field-product {{ grid-column: 5 / 7; }}
    label {{
      font-size: 12px;
      color: var(--muted);
      display: block;
      margin-bottom: 5px;
      font-weight: 500;
    }}
    input {{
      width: 100%;
      padding: 10px 11px;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
      color: var(--ink);
      outline: none;
      transition: border-color 140ms ease, box-shadow 140ms ease;
    }}
    select {{
      width: 100%;
      padding: 10px 11px;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
      color: var(--ink);
      outline: none;
    }}
    input:focus {{
      border-color: var(--accent-2);
      box-shadow: 0 0 0 3px rgba(15, 111, 101, 0.15);
    }}
    select:focus {{
      border-color: var(--accent-2);
      box-shadow: 0 0 0 3px rgba(15, 111, 101, 0.15);
    }}
    .impact-details {{
      position: relative;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
      overflow: visible;
    }}
    .impact-details > summary {{
      list-style: none;
      cursor: pointer;
      padding: 10px 11px;
      font-size: 14px;
      user-select: none;
    }}
    .impact-details > summary::-webkit-details-marker {{ display: none; }}
    .impact-details[open] > summary {{
      border-bottom: 1px solid var(--line);
      background: #f9fbfa;
    }}
    .impact-list {{
      position: absolute;
      top: calc(100% + 6px);
      left: 0;
      width: 320px;
      max-width: min(86vw, 360px);
      z-index: 60;
      max-height: 240px;
      overflow: auto;
      padding: 8px;
      display: grid;
      gap: 4px;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fff;
      box-shadow: 0 12px 28px rgba(25, 39, 45, 0.18);
    }}
    .impact-option {{
      display: grid;
      grid-template-columns: 16px minmax(0, 1fr);
      align-items: center;
      gap: 9px;
      font-size: 13px;
      min-height: 34px;
      padding: 6px 8px;
      border-radius: 6px;
      line-height: 1.2;
    }}
    .impact-option input[type="checkbox"] {{
      width: 14px;
      height: 14px;
      margin: 0;
      accent-color: #0f6f65;
    }}
    .impact-option span {{
      display: inline-block;
      min-width: 0;
      font-weight: 500;
      color: #2f3f45;
    }}
    .impact-option:hover {{ background: #f6faf9; }}
    .impact-selected {{
      margin-top: 6px;
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
    }}
    .impact-chip {{
      border: 1px solid #cfddd9;
      background: #f0faf7;
      color: #1a5852;
      border-radius: 999px;
      padding: 2px 8px;
      font-size: 11px;
      white-space: nowrap;
    }}
    .muted-chip {{
      border-color: #d6d9d7;
      background: #f6f7f6;
      color: #6b7471;
    }}
    button {{
      width: 100%;
      padding: 11px 14px;
      border: 0;
      border-radius: 10px;
      color: #fff;
      background: linear-gradient(135deg, var(--accent-2), #168173);
      font-weight: 700;
      cursor: pointer;
      transition: transform 100ms ease, filter 160ms ease;
    }}
    button:hover {{ filter: brightness(1.04); }}
    button:active {{ transform: translateY(1px); }}
    .search-btn {{ align-self: end; }}
    .field-cvss, .field-limit {{ max-width: 110px; }}
    .field-cpe-missing {{
      display: flex;
      align-items: end;
      padding-bottom: 8px;
    }}
    .field-cpe-missing label {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      margin: 0;
      font-size: 13px;
      color: #2f3f45;
      cursor: pointer;
    }}
    .field-cpe-missing input[type="checkbox"] {{
      width: 15px;
      height: 15px;
      margin: 0;
      accent-color: #0f6f65;
    }}
    .actions-bar {{
      display: flex;
      gap: 8px;
      align-items: end;
      align-self: end;
      grid-column: 1 / 7;
    }}
    .secondary-btn {{
      width: auto;
      background: #f1f4f3;
      color: #1f4048;
      border: 1px solid #ccd6d3;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 10px 12px;
      border-radius: 10px;
      font-size: 13px;
      font-weight: 600;
      min-height: 42px;
      cursor: pointer;
      white-space: nowrap;
    }}
    .secondary-btn:hover {{ background: #e8edeb; }}
    .export-dialog {{
      border: 1px solid #cfdad6;
      border-radius: 14px;
      padding: 0;
      width: min(92vw, 420px);
      box-shadow: 0 18px 42px rgba(18, 58, 68, 0.26);
    }}
    .export-dialog::backdrop {{
      background: rgba(21, 40, 48, 0.42);
      backdrop-filter: blur(2px);
    }}
    .export-dialog-body {{
      display: block;
      padding: 18px;
      border: 0;
      background: transparent;
      min-width: 0;
      max-width: 100%;
      overflow: visible;
    }}
    .export-dialog h3 {{
      margin: 0 0 8px;
      font-size: 17px;
      color: #123a44;
    }}
    .export-dialog p {{
      margin: 0;
      color: #4b5f64;
      font-size: 13px;
      line-height: 1.45;
    }}
    .export-dialog-actions {{
      margin-top: 16px;
      display: flex;
      gap: 8px;
      justify-content: flex-end;
      flex-wrap: nowrap;
    }}
    .dialog-btn {{
      width: auto;
      min-height: 36px;
      padding: 8px 11px;
      font-size: 12px;
      border-radius: 9px;
    }}
    .dialog-btn.page {{
      background: #eef3f2;
      color: #21464f;
      border: 1px solid #c6d6d1;
    }}
    .dialog-btn.cancel {{
      background: #f8f4f0;
      color: #6e4f3d;
      border: 1px solid #dfc9b8;
    }}
    .meta {{
      margin: 0;
      padding: 12px 16px 4px;
      color: var(--muted);
      font-size: 13px;
      text-align: right;
    }}
    .pager {{
      margin: 8px 16px 0;
      display: flex;
      justify-content: flex-end;
      align-items: center;
      gap: 6px;
      flex-wrap: wrap;
    }}
    .page-link {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 32px;
      padding: 5px 9px;
      border: 1px solid #ccd6d3;
      border-radius: 8px;
      background: #fff;
      color: #1f4048;
      text-decoration: none;
      font-size: 12px;
      font-weight: 600;
    }}
    .page-link.current {{
      background: #e5f3ef;
      border-color: #9ec2b8;
      color: #0f6f65;
    }}
    .page-link.disabled {{
      background: #f1f3f2;
      border-color: #d9dfdc;
      color: #9aa4a1;
    }}
    .error {{
      margin: 8px 16px 0;
      padding: 10px 11px;
      border: 1px solid rgba(194, 72, 46, 0.35);
      background: rgba(194, 72, 46, 0.08);
      border-radius: 10px;
      color: #8f2917;
      font-size: 13px;
    }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
    thead th {{
      position: sticky;
      top: 0;
      background: #f2eee4;
      text-align: center;
      vertical-align: middle;
      font-size: 12px;
      letter-spacing: 0.4px;
      text-transform: uppercase;
      color: var(--muted);
    }}
    thead th a {{
      color: inherit;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 5px;
      width: 100%;
    }}
    thead th a:hover {{ color: #1e5a53; }}
    .sort-mark {{ font-size: 11px; opacity: 0.9; }}
    th, td {{
      border-top: 1px solid var(--line);
      padding: 10px 12px;
      vertical-align: top;
      font-size: 14px;
    }}
    tbody tr:hover {{ background: #fff8ec; }}
    .id {{ width: 190px; white-space: nowrap; font-weight: 700; color: #123a44; }}
    .score {{ width: 92px; white-space: nowrap; }}
    .actions {{ width: 120px; white-space: nowrap; }}
    .cvss-chip {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 84px;
      padding: 3px 9px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 700;
      border: 1px solid transparent;
    }}
    .cvss-critical {{
      color: #9f1f1f;
      background: #fde8e8;
      border-color: #efb6b6;
    }}
    .cvss-high {{
      color: #9a4a00;
      background: #fff1e4;
      border-color: #f0c79c;
    }}
    .cvss-medium {{
      color: #7b6400;
      background: #fff8d8;
      border-color: #ead88a;
    }}
    .cvss-low {{
      color: #4b4f55;
      background: #eef0f3;
      border-color: #d3d8de;
    }}
    .cvss-none {{
      color: #8f98a3;
      background: #1b1f24;
      border-color: #2f3842;
    }}
    td.cpe, td.actions {{
      vertical-align: middle;
      text-align: left;
    }}
    .copy-btn {{
      width: auto;
      font-size: 11px;
      padding: 5px 8px;
      border-radius: 999px;
      border: 1px solid #bfd1cb;
      background: #ecf7f4;
      color: #164f49;
      margin-right: 6px;
      margin-top: 3px;
      cursor: pointer;
    }}
    .copy-btn.alt {{
      background: #eef3f7;
      border-color: #c5d0d9;
      color: #2d4658;
    }}
    .copy-btn:hover {{ filter: brightness(0.98); }}
    .desc {{
      color: #334b53;
      font-weight: 500;
    }}
    .view-btn {{
      background: #f7f2ff;
      border-color: #d2c2ee;
      color: #49356a;
    }}
    .desc-drawer {{
      position: fixed;
      top: 0;
      right: 0;
      width: min(720px, 94vw);
      height: 100vh;
      background: #fffdf8;
      border-left: 1px solid var(--line);
      box-shadow: -10px 0 24px rgba(26, 36, 42, 0.18);
      z-index: 300;
      transform: translateX(102%);
      transition: transform 170ms ease;
      display: flex;
      flex-direction: column;
    }}
    .desc-drawer.open {{
      transform: translateX(0);
    }}
    .desc-drawer-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 8px;
      padding: 14px 16px;
      border-bottom: 1px solid var(--line);
      background: #f7faf8;
    }}
    .desc-drawer-title {{
      margin: 0;
      font-size: 15px;
      font-weight: 700;
      color: #1f343b;
    }}
    .desc-drawer-body {{
      padding: 14px 16px;
      overflow: auto;
      line-height: 1.55;
      white-space: pre-wrap;
      color: #2f3f45;
      font-size: 14px;
    }}
    .desc-drawer-close {{
      width: auto;
      min-height: 34px;
      border: 1px solid #ced9d5;
      border-radius: 8px;
      padding: 6px 10px;
      background: #fff;
      color: #254149;
      font-size: 12px;
      font-weight: 700;
      cursor: pointer;
    }}
    .cpe-wrap {{
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      align-items: center;
      justify-content: flex-start;
    }}
    .cpe-chip {{
      display: inline-block;
      border: 1px solid #cfddd9;
      background: #f0faf7;
      color: #1a5852;
      padding: 3px 7px;
      border-radius: 999px;
      font-size: 12px;
      white-space: normal;
      overflow-wrap: break-word;
      word-break: normal;
      max-width: 100%;
    }}
    @keyframes rise {{
      from {{ opacity: 0; transform: translateY(10px); }}
      to {{ opacity: 1; transform: translateY(0); }}
    }}
    @media (max-width: 900px) {{
      .wrap {{ width: min(1120px, 96vw); margin-top: 16px; }}
      .checkpoint-badge {{ width: 100%; margin-left: 0; text-align: left; }}
      .panel > form {{ grid-template-columns: 1fr 1fr; }}
      .field-lastmod-start,
      .field-lastmod-end,
      .field-vendor,
      .field-product,
      .field-keyword,
      .field-cvss,
      .field-impact,
      .field-cpe-missing,
      .field-limit {{
        grid-column: auto;
      }}
      .search-btn {{ grid-column: 1 / -1; }}
      .actions-bar {{ grid-column: 1 / -1; justify-content: flex-start; }}
      .impact-list {{ width: min(92vw, 360px); }}
      .detail-body {{
        position: static;
        width: 100%;
        min-width: 0;
        max-height: none;
        margin-top: 8px;
      }}
      table, thead, tbody, th, td, tr {{ display: block; }}
      thead {{ display: none; }}
      td {{
        border-top: 0;
        padding: 6px 12px;
      }}
      tbody tr {{
        padding: 8px 0;
        border-top: 1px solid var(--line);
      }}
      td.id::before {{ content: "CVE ID"; display: block; font-size: 12px; color: var(--muted); }}
      td.score::before {{ content: "CVSS"; display: block; font-size: 12px; color: var(--muted); }}
      td.lastmod::before {{ content: "Last Modified"; display: block; font-size: 12px; color: var(--muted); }}
      td.vtype::before {{ content: "Type"; display: block; font-size: 12px; color: var(--muted); }}
      td.desc::before {{ content: "Description"; display: block; font-size: 12px; color: var(--muted); }}
      td.cpe::before {{ content: "CPE"; display: block; font-size: 12px; color: var(--muted); }}
      td.actions::before {{ content: "Actions"; display: block; font-size: 12px; color: var(--muted); }}
    }}
  </style>
</head>
  <body>
  <main class="wrap">
    {menu_html}
    <section class="hero">
      <div class="hero-copy">
        <h1>CVE Explorer</h1>
        <p class="sub">Search by vendor/product and Last Modified range using normalized CPE mappings.</p>
      </div>
      <div class="checkpoint-badge">
        <strong>최종 증분 수집 시각</strong>
        {escape(checkpoint_text)}
      </div>
    </section>
    <section class="panel">
      <form method="get">
        <input id="user_profile" type="hidden" name="user_profile" value="{escape(user_profile)}">
        <input type="hidden" name="impact_type_present" value="1">
        <input type="hidden" name="cpe_missing_only_present" value="1">
        <input type="hidden" name="last_modified_present" value="1">
        <input type="hidden" name="cpe_object_present" value="1">
        <div class="field-lastmod-start">
          <label for="last_modified_start_date">Last Modified Start</label>
          <div class="datetime-parts">
            <input id="last_modified_start_date" name="last_modified_start_date" type="date" value="{escape(last_modified_start_date_raw)}">
            <input id="last_modified_start_time" name="last_modified_start_time" type="text" class="datetime-time" inputmode="numeric" placeholder="HH:MM" pattern="\\d{{1,2}}:\\d{{2}}" title="HH:MM (24-hour)" value="{escape(last_modified_start_time_raw)}">
          </div>
        </div>
        <div class="field-lastmod-end">
          <label for="last_modified_end_date">Last Modified End</label>
          <div class="datetime-parts">
            <input id="last_modified_end_date" name="last_modified_end_date" type="date" value="{escape(last_modified_end_date_raw)}">
            <input id="last_modified_end_time" name="last_modified_end_time" type="text" class="datetime-time" inputmode="numeric" placeholder="HH:MM" pattern="\\d{{1,2}}:\\d{{2}}" title="HH:MM (24-hour)" value="{escape(last_modified_end_time_raw)}">
          </div>
        </div>
        <div class="field-cvss">
          <label for="min_cvss">Min CVSS</label>
          <input id="min_cvss" name="min_cvss" type="number" min="0" max="10" step="0.1" value="{escape(str(min_cvss))}">
        </div>
        <div class="field-impact">
          <label for="impact_type">Impact Type</label>
          <details class="impact-details">
            <summary>{escape(impact_summary)}</summary>
            <div class="impact-list">
              {impact_options_html}
            </div>
          </details>
          <div class="impact-selected">{impact_selected_html}</div>
        </div>
        <div class="field-limit">
          <label for="limit">Limit (1-500)</label>
          <input id="limit" name="limit" type="number" min="1" max="500" step="1" value="{escape(str(limit))}">
        </div>
        <div class="field-cpe-missing">
          <label for="cpe_missing_only">
            <input id="cpe_missing_only" name="cpe_missing_only" type="checkbox" value="1" {cpe_missing_checked}>
            CPE missing only
          </label>
        </div>
        <div class="search-btn">
          <button type="submit">Search CVEs</button>
        </div>
        <div class="field-keyword">
          <label for="keyword">Keyword (comma = OR; description/vendor/product)</label>
          <input id="keyword" name="keyword" value="{escape(keyword)}" placeholder="e.g. ssl, auth bypass, endpoint">
        </div>
        <div class="field-vendor">
          <label for="vendor">Vendor</label>
          <input id="vendor" name="vendor" value="{escape(vendor)}" placeholder="e.g. ivanti">
        </div>
        <div class="field-product">
          <label for="product">Product (comma = OR)</label>
          <input id="product" name="product" value="{escape(product)}" placeholder="e.g. endpoint_manager_mobile, pulse_connect_secure">
        </div>
        <div class="field-impact">
          <label for="cpe_object">CPE Object (on/off)</label>
          <details class="impact-details">
            <summary>{escape(cpe_object_summary)}</summary>
            <div class="impact-list">
              {cpe_object_options_html}
            </div>
          </details>
          <div class="impact-selected">{cpe_object_selected_html}</div>
        </div>
        <input type="hidden" name="sort_key" value="{escape(sort_key)}">
        <div class="actions-bar">
          <a class="secondary-btn" href="/">Reset Filters</a>
          <button id="export-xlsx-btn" type="button" class="secondary-btn">Export Excel</button>
          <button id="share-url-btn" type="button" class="secondary-btn">Share URL</button>
        </div>
      </form>
      <p class="meta">Results: {total_count} total (showing {len(rows)})</p>
      {pager_html}
      {error_html}
      <table>
        <thead>
          <tr>
            <th>CVE ID</th>
            <th><a href="{escape(cvss_sort_href)}">CVSS <span class="sort-mark">{escape(cvss_sort_marker)}</span></a></th>
            <th><a href="{escape(last_modified_sort_href)}">Last Modified <span class="sort-mark">{escape(last_modified_sort_marker)}</span></a></th>
            <th>Type</th>
            <th>Description</th>
            <th>CPE</th>
            <th>Actions</th>
          </tr>
        </thead>
        <tbody>
          {rows_html}
        </tbody>
      </table>
    </section>
  </main>
  <aside id="desc-drawer" class="desc-drawer" aria-hidden="true">
    <div class="desc-drawer-header">
      <h3 id="desc-drawer-title" class="desc-drawer-title">Description</h3>
      <button id="desc-drawer-close" type="button" class="desc-drawer-close">닫기</button>
    </div>
    <div id="desc-drawer-body" class="desc-drawer-body"></div>
  </aside>
  <dialog id="export-dialog" class="export-dialog">
    <form method="dialog" class="export-dialog-body">
      <h3>엑셀 내보내기 범위 선택</h3>
      <p>현재 검색 조건 결과를 엑셀로 다운로드합니다. 전체 결과 또는 현재 페이지만 선택하세요.</p>
      <div class="export-dialog-actions">
        <button type="button" class="dialog-btn" data-export-scope="all">전체 결과</button>
        <button type="button" class="dialog-btn page" data-export-scope="page">현재 페이지만</button>
        <button type="button" class="dialog-btn cancel" data-export-scope="cancel">취소</button>
      </div>
    </form>
  </dialog>
</body>
<script>
  (() => {{
    const form = document.querySelector("form");
    const impactDetails = document.querySelector(".impact-details");
    const shareButton = document.querySelector("#share-url-btn");
    const exportButton = document.querySelector("#export-xlsx-btn");
    const exportDialog = document.querySelector("#export-dialog");
    const lastModifiedStartDateInput = document.querySelector("#last_modified_start_date");
    const lastModifiedEndDateInput = document.querySelector("#last_modified_end_date");
    const copyButtons = document.querySelectorAll(".copy-btn");
    const descDrawer = document.querySelector("#desc-drawer");
    const descDrawerBody = document.querySelector("#desc-drawer-body");
    const descDrawerTitle = document.querySelector("#desc-drawer-title");
    const descDrawerClose = document.querySelector("#desc-drawer-close");

    const syncLastModifiedDateRange = () => {{
      if (!lastModifiedStartDateInput || !lastModifiedEndDateInput) {{
        return;
      }}
      const startValue = (lastModifiedStartDateInput.value || "").trim();
      if (startValue) {{
        lastModifiedEndDateInput.min = startValue;
        lastModifiedEndDateInput.title = `End date must be on or after ${{startValue}}`;
      }} else {{
        lastModifiedEndDateInput.removeAttribute("min");
        lastModifiedEndDateInput.title = "";
      }}
      const endValue = (lastModifiedEndDateInput.value || "").trim();
      if (startValue && endValue && endValue < startValue) {{
        lastModifiedEndDateInput.value = startValue;
      }}
    }};

    const openExportDialog = () => {{
      if (!exportDialog || typeof exportDialog.showModal !== "function") {{
        return Promise.resolve(window.confirm("전체 결과를 내보낼까요?\\n확인: 전체 결과\\n취소: 현재 페이지만") ? "all" : "page");
      }}
      return new Promise((resolve) => {{
        const buttons = exportDialog.querySelectorAll("[data-export-scope]");
        const onSelect = (event) => {{
          const target = event.currentTarget;
          const scope = target?.dataset?.exportScope || "cancel";
          cleanup();
          exportDialog.close();
          resolve(scope);
        }};
        const onClose = () => {{
          cleanup();
          resolve("cancel");
        }};
        const cleanup = () => {{
          buttons.forEach((btn) => btn.removeEventListener("click", onSelect));
          exportDialog.removeEventListener("close", onClose);
        }};
        buttons.forEach((btn) => btn.addEventListener("click", onSelect));
        exportDialog.addEventListener("close", onClose, {{ once: true }});
        exportDialog.showModal();
      }});
    }};

    if (impactDetails) {{
      document.addEventListener("click", (event) => {{
        if (!impactDetails.open) return;
        if (impactDetails.contains(event.target)) return;
        impactDetails.open = false;
      }});

      document.addEventListener("keydown", (event) => {{
        if (event.key === "Escape" && impactDetails.open) {{
          impactDetails.open = false;
        }}
      }});
    }}

    if (shareButton) {{
      shareButton.addEventListener("click", async () => {{
        const url = window.location.href;
        try {{
          await navigator.clipboard.writeText(url);
          shareButton.textContent = "Copied URL";
          setTimeout(() => {{ shareButton.textContent = "Share URL"; }}, 1200);
        }} catch (_) {{
          window.prompt("Copy URL:", url);
        }}
      }});
    }}

    if (lastModifiedStartDateInput && lastModifiedEndDateInput) {{
      syncLastModifiedDateRange();
      lastModifiedStartDateInput.addEventListener("change", syncLastModifiedDateRange);
      lastModifiedStartDateInput.addEventListener("input", syncLastModifiedDateRange);
      lastModifiedEndDateInput.addEventListener("focus", syncLastModifiedDateRange);
      lastModifiedEndDateInput.addEventListener("change", syncLastModifiedDateRange);
    }}

    if (exportButton) {{
      exportButton.addEventListener("click", async () => {{
        const exportScope = await openExportDialog();
        if (exportScope === "cancel") {{
          return;
        }}
        const params = new URLSearchParams(window.location.search);
        if (form) {{
          const formData = new FormData(form);
          for (const [key, value] of formData.entries()) {{
            params.set(key, String(value));
          }}
          if (!formData.has("cpe_missing_only")) {{
            params.delete("cpe_missing_only");
          }}
          if (!formData.has("impact_type")) {{
            params.delete("impact_type");
          }}
          if (!formData.has("cpe_object")) {{
            params.delete("cpe_object");
          }}
        }}
        params.set("export_scope", exportScope);
        window.location.href = `/export.xlsx?${{params.toString()}}`;
      }});
    }}

    copyButtons.forEach((btn) => {{
      btn.addEventListener("click", async () => {{
        if (btn.classList.contains("view-btn")) {{
          const cve = btn.dataset.cve || "CVE";
          const desc = btn.dataset.desc || "";
          if (descDrawer && descDrawerBody && descDrawerTitle) {{
            descDrawerTitle.textContent = cve;
            descDrawerBody.textContent = desc;
            descDrawer.classList.add("open");
            descDrawer.setAttribute("aria-hidden", "false");
          }}
          return;
        }}
        const text = btn.dataset.copy || "";
        try {{
          await navigator.clipboard.writeText(text);
          const before = btn.textContent;
          btn.textContent = "Copied";
          setTimeout(() => {{ btn.textContent = before; }}, 900);
        }} catch (_) {{
          window.prompt("Copy value:", text);
        }}
      }});
    }});
    descDrawerClose?.addEventListener("click", () => {{
      if (!descDrawer) return;
      descDrawer.classList.remove("open");
      descDrawer.setAttribute("aria-hidden", "true");
    }});
    document.addEventListener("keydown", (event) => {{
      if (event.key === "Escape" && descDrawer?.classList.contains("open")) {{
        descDrawer.classList.remove("open");
        descDrawer.setAttribute("aria-hidden", "true");
      }}
    }});
  }})();
</script>
</html>
"""


@app.route("/daily", methods=["GET", "POST"])
def daily_review() -> str | object:
    user_profile = _normalize_user_profile(request.values.get("user_profile"))
    app_settings: Settings | None = None
    profile_defaults = dict(DEFAULT_PROFILE_SETTINGS)
    error_text = ""
    notice_text = (request.args.get("notice") or "").strip()
    highlight_cve_id = (request.args.get("highlight_cve") or "").strip()
    try:
        app_settings = load_settings(".env")
        profile_defaults = fetch_profile_settings(app_settings, user_profile)
    except Exception as exc:  # pragma: no cover
        error_text = f"설정 로딩 실패: {exc}"
    active_presets: list[dict[str, object]] = []
    if app_settings is not None:
        try:
            active_presets = [item for item in fetch_profile_presets(app_settings, user_profile) if item["is_enabled"]]
        except Exception:
            active_presets = []

    now_local = datetime.now().replace(second=0, microsecond=0)
    window_days_raw = request.values.get("window_days") or str(profile_defaults["daily_review_window_days"])
    try:
        window_days = max(1, min(int(window_days_raw), 30))
    except ValueError:
        window_days = int(profile_defaults["daily_review_window_days"])
    review_limit_raw = request.values.get("review_limit") or str(profile_defaults["daily_review_limit"])
    try:
        review_limit = max(1, min(int(review_limit_raw), 1000))
    except ValueError:
        review_limit = int(profile_defaults["daily_review_limit"])

    period_mode = (request.values.get("period_mode") or "previous_day").strip().lower()
    if period_mode not in {"previous_day", "last24h"}:
        period_mode = "previous_day"
    status_filter = (request.values.get("status_filter") or "pending").strip().lower()
    if status_filter not in {"all", "pending", "reviewed", "ignored"}:
        status_filter = "pending"

    if period_mode == "previous_day":
        end_dt = now_local.replace(hour=0, minute=0)
        start_dt = end_dt - timedelta(days=window_days)
        review_date = (end_dt - timedelta(days=1)).date().isoformat()
        period_label = f"{start_dt.strftime('%Y-%m-%d %H:%M')} ~ {end_dt.strftime('%Y-%m-%d %H:%M')}"
    else:
        end_dt = now_local
        start_dt = now_local - timedelta(hours=24 * window_days)
        review_date = end_dt.date().isoformat()
        period_label = f"{start_dt.strftime('%Y-%m-%d %H:%M')} ~ {end_dt.strftime('%Y-%m-%d %H:%M')}"

    if request.method == "POST" and not error_text:
        action = (request.form.get("action") or "row_update").strip().lower()
        redirect_query = {
            "user_profile": user_profile,
            "period_mode": period_mode,
            "window_days": str(window_days),
            "review_limit": str(review_limit),
            "status_filter": status_filter,
        }
        if action == "bulk_update":
            selected_cve_ids = [value.strip() for value in request.form.getlist("selected_cve_id") if value.strip()]
            bulk_status = (request.form.get("bulk_status") or "pending").strip().lower()
            bulk_note = (request.form.get("bulk_note") or "").strip()
            if not selected_cve_ids:
                error_text = "일괄 변경할 CVE를 선택하세요."
            else:
                try:
                    previous_state = fetch_daily_review_map(app_settings, user_profile, review_date)
                    _last_bulk_action_cache[f"{user_profile}:{review_date}"] = {
                        "items": [
                            {
                                "cve_id": cve_id,
                                "status": previous_state.get(cve_id, {}).get("status", "pending"),
                                "note": previous_state.get(cve_id, {}).get("note", ""),
                            }
                            for cve_id in selected_cve_ids
                        ]
                    }
                    for cve_id in selected_cve_ids:
                        upsert_daily_review_item(app_settings, user_profile, review_date, cve_id, bulk_status, bulk_note)
                    redirect_query["notice"] = f"{len(selected_cve_ids)}건 상태가 일괄 저장되었습니다."
                    return redirect("/daily?" + urlencode(redirect_query))
                except Exception as exc:  # pragma: no cover
                    error_text = f"일괄 상태 저장 실패: {exc}"
        elif action == "undo_bulk":
            last_bulk = _last_bulk_action_cache.get(f"{user_profile}:{review_date}", {})
            items = list(last_bulk.get("items", []))
            if not items:
                error_text = "되돌릴 최근 일괄 변경이 없습니다."
            else:
                try:
                    for item in items:
                        upsert_daily_review_item(
                            app_settings,
                            user_profile,
                            review_date,
                            str(item.get("cve_id", "")),
                            str(item.get("status", "pending")),
                            str(item.get("note", "")),
                        )
                    _last_bulk_action_cache.pop(f"{user_profile}:{review_date}", None)
                    redirect_query["notice"] = f"{len(items)}건 일괄 변경을 되돌렸습니다."
                    return redirect("/daily?" + urlencode(redirect_query))
                except Exception as exc:  # pragma: no cover
                    error_text = f"일괄 상태 되돌리기 실패: {exc}"
        else:
            cve_id = (request.form.get("cve_id") or "").strip()
            status = (request.form.get("status") or "pending").strip().lower()
            note = (request.form.get("note") or "").strip()
            if not cve_id:
                error_text = "상태를 저장할 CVE ID가 없습니다."
            else:
                try:
                    upsert_daily_review_item(app_settings, user_profile, review_date, cve_id, status, note)
                    redirect_query["notice"] = f"{cve_id} 상태가 저장되었습니다."
                    redirect_query["highlight_cve"] = cve_id
                    return redirect("/daily?" + urlencode(redirect_query))
                except Exception as exc:  # pragma: no cover
                    error_text = f"상태 저장 실패: {exc}"

    rows: list[dict[str, object]] = []
    total_count = 0
    review_map: dict[str, dict[str, str]] = {}
    matched_preset_map: dict[str, list[str]] = {}
    if not error_text:
        try:
            if active_presets:
                merged_by_cve: dict[str, dict[str, object]] = {}
                for preset in active_presets:
                    preset_name = str(preset["preset_name"])
                    rule = dict(preset["rule"])
                    preset_rows, _ = fetch_cves_from_db(
                        app_settings,
                        str(rule["product"]) or None,
                        str(rule["vendor"]) or None,
                        str(rule["keyword"]) or None,
                        list(rule["impact_type"]) or None,
                        float(rule["min_cvss"]),
                        review_limit,
                        offset=0,
                        sort_by="last_modified",
                        sort_order="desc",
                        last_modified_start=start_dt,
                        last_modified_end=end_dt,
                        cpe_missing_only=bool(rule["cpe_missing_only"]),
                        cpe_objects=list(rule["cpe_objects_catalog"]) or None,
                        include_total_count=False,
                    )
                    for row in preset_rows:
                        cve_id = str(row.get("id", ""))
                        if not cve_id:
                            continue
                        if cve_id not in merged_by_cve:
                            merged_by_cve[cve_id] = row
                        matched_preset_map.setdefault(cve_id, [])
                        if preset_name not in matched_preset_map[cve_id]:
                            matched_preset_map[cve_id].append(preset_name)
                rows = list(merged_by_cve.values())
                rows.sort(
                    key=lambda row: (
                        row.get("last_modified_at") or datetime.min,
                        float(row.get("cvss_score") or 0.0),
                    ),
                    reverse=True,
                )
                total_count = len(rows)
                rows = rows[:review_limit]
            else:
                rows, total_count = fetch_cves_from_db(
                    app_settings,
                    str(profile_defaults["product"]) or None,
                    str(profile_defaults["vendor"]) or None,
                    str(profile_defaults["keyword"]) or None,
                    list(profile_defaults["impact_type"]) or None,
                    float(profile_defaults["min_cvss"]),
                    review_limit,
                    offset=0,
                    sort_by="last_modified",
                    sort_order="desc",
                    last_modified_start=start_dt,
                    last_modified_end=end_dt,
                    cpe_missing_only=bool(profile_defaults["cpe_missing_only"]),
                    cpe_objects=list(profile_defaults["cpe_objects_catalog"]) or None,
                    include_total_count=True,
                )
            review_map = fetch_daily_review_map(app_settings, user_profile, review_date)
        except Exception as exc:  # pragma: no cover
            error_text = str(exc)

    status_summary = {"pending": 0, "reviewed": 0, "ignored": 0}
    filtered_count = 0
    row_chunks: list[str] = []
    for row in rows:
        cve_id_raw = str(row.get("id", "UNKNOWN"))
        cve_id = escape(cve_id_raw)
        state = review_map.get(cve_id_raw, {"status": "pending", "note": ""})
        current_status = state.get("status", "pending")
        if current_status not in status_summary:
            current_status = "pending"
        status_summary[current_status] += 1
        if status_filter != "all" and current_status != status_filter:
            continue
        filtered_count += 1
        current_note = escape(state.get("note", ""))
        score_label, score_class = format_cvss_badge(row.get("cvss_score"))
        status_badge_class = "review-badge pending"
        status_badge_text = "미검토"
        if current_status == "reviewed":
            status_badge_class = "review-badge reviewed"
            status_badge_text = "검토완료"
        elif current_status == "ignored":
            status_badge_class = "review-badge ignored"
            status_badge_text = "제외"
        row_chunks.append(
            "<tr>"
            f"<td class='sticky-col sticky-left'><input type='checkbox' name='selected_cve_id' value='{cve_id}' form='bulk-form' class='bulk-cve-check'></td>"
            f"<td class='id'>{cve_id}</td>"
            f"<td class='score'><span class='cvss-chip {escape(score_class)}'>{escape(score_label)}</span></td>"
            f"<td>{escape(str(row.get('vuln_type', 'Other')))}</td>"
            f"<td>{escape(format_last_modified(row.get('last_modified_at', 'N/A')))}</td>"
            f"<td>{escape(shorten(str(row.get('description', '')), 120))}</td>"
            f"<td>{escape(', '.join(matched_preset_map.get(cve_id_raw, [])) or '-')}</td>"
            f"<td class='sticky-col sticky-right {'row-highlight' if cve_id_raw == highlight_cve_id else ''}'>"
            f"<div class='{status_badge_class}'>{status_badge_text}</div>"
            "<form method='post' class='review-form'>"
            f"<input type='hidden' name='user_profile' value='{escape(user_profile)}'>"
            f"<input type='hidden' name='period_mode' value='{escape(period_mode)}'>"
            f"<input type='hidden' name='window_days' value='{window_days}'>"
            f"<input type='hidden' name='review_limit' value='{review_limit}'>"
            f"<input type='hidden' name='status_filter' value='{escape(status_filter)}'>"
            "<input type='hidden' name='action' value='row_update'>"
            f"<input type='hidden' name='cve_id' value='{cve_id}'>"
            "<select name='status'>"
            f"<option value='pending' {'selected' if current_status == 'pending' else ''}>미검토</option>"
            f"<option value='reviewed' {'selected' if current_status == 'reviewed' else ''}>검토완료</option>"
            f"<option value='ignored' {'selected' if current_status == 'ignored' else ''}>제외</option>"
            "</select>"
            f"<input name='note' value='{current_note}' placeholder='메모 (선택)'>"
            "<button type='submit'>저장</button>"
            "</form>"
            "</td>"
            "</tr>"
        )
    if not row_chunks:
        row_chunks.append("<tr><td colspan='8'>대상 없음</td></tr>")

    info_lines = [
        f"기간: {period_label}",
        f"필터: vendor={profile_defaults['vendor'] or '-'}, product={profile_defaults['product'] or '-'}, keyword={profile_defaults['keyword'] or '-'}",
        f"필터: impact={', '.join(profile_defaults['impact_type']) if profile_defaults['impact_type'] else '-'}, cpe_objects={len(profile_defaults['cpe_objects_catalog'])}",
        f"검토상태: 미검토 {status_summary['pending']} / 검토완료 {status_summary['reviewed']} / 제외 {status_summary['ignored']}",
        f"현재 표시 필터: {status_filter} (표시 {filtered_count}건)",
        f"활성 프리셋: {', '.join(item['preset_name'] for item in active_presets) if active_presets else '없음(프로필 기본 규칙 사용)'}",
    ]
    has_undo_bulk = bool(_last_bulk_action_cache.get(f"{user_profile}:{review_date}", {}).get("items"))
    quick_status_links = "".join(
        f"<a class='quick-chip {'active' if status_filter == key else ''}' href='/daily?{urlencode({'user_profile': user_profile, 'period_mode': period_mode, 'window_days': str(window_days), 'review_limit': str(review_limit), 'status_filter': key})}'>{label}</a>"
        for key, label in [("pending", "미검토"), ("reviewed", "검토완료"), ("ignored", "제외"), ("all", "전체")]
    )
    quick_period_links = "".join(
        f"<a class='quick-chip {'active' if period_mode == key else ''}' href='/daily?{urlencode({'user_profile': user_profile, 'period_mode': key, 'window_days': str(window_days), 'review_limit': str(review_limit), 'status_filter': status_filter})}'>{label}</a>"
        for key, label in [("previous_day", "전일 마감"), ("last24h", "최근 24h")]
    )
    daily_export_href = (
        "/daily_export.xlsx?"
        + urlencode(
            {
                "user_profile": user_profile,
                "period_mode": period_mode,
                "window_days": str(window_days),
                "review_limit": str(review_limit),
                "status_filter": status_filter,
            }
        )
    )
    notice_html = f"<p class='ok-msg'>{escape(notice_text)}</p>" if notice_text else ""
    error_html = f"<p class='error-msg'>{escape(error_text)}</p>" if error_text else ""
    menu_html = _build_menu_html("daily", user_profile=user_profile)

    return f"""
<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CVE Daily Review</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;700&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg: #f7f4ee; --panel: #fffdf8; --ink: #1e2b31; --muted: #5e6c73; --line: #d7d5cc; --accent-2: #0f6f65;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Space Grotesk", "Pretendard", sans-serif; color: var(--ink); background: var(--bg); }}
    .wrap {{ width: min(1600px, 90vw); margin: 34px auto 54px; }}
    .top-menu {{ display:flex; gap:10px; margin-bottom:12px; }}
    .menu-link {{ text-decoration:none; color:var(--ink); border:1px solid var(--line); background:#fff8ed; border-radius:999px; padding:8px 14px; font-size:13px; font-weight:700; }}
    .menu-link.active {{ color:#fff; background:var(--accent-2); border-color:var(--accent-2); }}
    .panel {{ border:1px solid var(--line); background:var(--panel); border-radius:16px; padding:14px; }}
    .profile-tabs {{ display:flex; gap:8px; margin-bottom:10px; }}
    .profile-tab {{ text-decoration:none; border:1px solid var(--line); color:var(--ink); padding:6px 12px; border-radius:999px; font-size:13px; font-weight:700; background:#fff; }}
    .profile-tab.active {{ color:#fff; background:var(--accent-2); border-color:var(--accent-2); }}
    .toolbar {{ display:flex; gap:8px; flex-wrap:wrap; align-items:flex-end; margin-bottom:10px; }}
    .toolbar label {{ font-size:12px; color:var(--muted); display:block; margin-bottom:4px; }}
    .toolbar select,.toolbar input {{ border:1px solid var(--line); border-radius:8px; padding:8px 10px; font:inherit; background:#fff; }}
    .toolbar button {{ border:1px solid var(--line); border-radius:8px; padding:8px 12px; font:inherit; font-weight:700; cursor:pointer; background:#fff; }}
    .toolbar-link {{
      text-decoration: none;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px 12px;
      font: inherit;
      font-weight: 700;
      background: #f7f8f8;
      color: #244149;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 38px;
    }}
    .quick-filters {{
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      margin: 0 0 10px;
    }}
    .quick-chip {{
      text-decoration: none;
      border: 1px solid #ccd6d3;
      border-radius: 999px;
      padding: 5px 10px;
      font-size: 12px;
      font-weight: 700;
      color: #28444c;
      background: #fff;
    }}
    .quick-chip.active {{
      color: #fff;
      background: var(--accent-2);
      border-color: var(--accent-2);
    }}
    .bulk-toolbar {{ margin: 0 0 10px; }}
    .meta {{ margin: 8px 0 10px; font-size: 13px; color: var(--muted); }}
    .ok-msg {{ color:#0c6d57; font-size:13px; font-weight:700; margin: 4px 0; }}
    .error-msg {{ color:#ad3427; font-size:13px; font-weight:700; margin: 4px 0; }}
    table {{ width:100%; border-collapse: collapse; }}
    th, td {{ border-top:1px solid var(--line); padding:8px 10px; vertical-align:top; font-size:13px; }}
    th {{ text-align:left; color:var(--muted); font-size:12px; text-transform:uppercase; }}
    .id {{ white-space:nowrap; font-weight:700; }}
    .review-form {{ display:grid; grid-template-columns: 120px 1fr auto; gap:6px; }}
    .review-form select,.review-form input,.review-form button {{ border:1px solid var(--line); border-radius:7px; padding:6px 8px; font:inherit; }}
    .review-form button {{ font-weight:700; background:#f0faf7; color:#1a5852; cursor:pointer; }}
    .cvss-chip {{ display:inline-flex; min-width:84px; justify-content:center; border-radius:999px; padding:2px 8px; font-size:12px; border:1px solid transparent; }}
    .cvss-critical {{ color:#9f1f1f; background:#fde8e8; border-color:#efb6b6; }}
    .cvss-high {{ color:#9a4a00; background:#fff1e4; border-color:#f0c79c; }}
    .cvss-medium {{ color:#7b6400; background:#fff8d8; border-color:#ead88a; }}
    .cvss-low {{ color:#4b4f55; background:#eef0f3; border-color:#d3d8de; }}
    .cvss-none {{ color:#8f98a3; background:#1b1f24; border-color:#2f3842; }}
    .table-wrap {{ overflow: auto; border: 1px solid var(--line); border-radius: 10px; }}
    .sticky-col {{
      position: sticky;
      background: #fffdf8;
      z-index: 3;
    }}
    .sticky-left {{ left: 0; min-width: 42px; }}
    .sticky-right {{ right: 0; min-width: 340px; box-shadow: -6px 0 8px rgba(20, 34, 40, 0.04); }}
    .review-badge {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 72px;
      padding: 2px 8px;
      border-radius: 999px;
      border: 1px solid transparent;
      font-size: 11px;
      font-weight: 700;
      margin-bottom: 6px;
    }}
    .review-badge.pending {{ background: #f1f3f5; color: #3f4950; border-color: #d4dade; }}
    .review-badge.reviewed {{ background: #e7f6ef; color: #0e6a4c; border-color: #9ecfb9; }}
    .review-badge.ignored {{ background: #fff0ea; color: #8d3a21; border-color: #efb8a3; }}
    .row-highlight {{
      animation: rowGlow 1.2s ease-out;
      box-shadow: inset 0 0 0 2px rgba(15, 111, 101, 0.22);
    }}
    @keyframes rowGlow {{
      0% {{ background: #e9f8f3; }}
      100% {{ background: #fffdf8; }}
    }}
  </style>
</head>
<body>
  <main class="wrap">
    {menu_html}
    <section class="panel">
      <h1 style="margin:0 0 8px;font-size:24px;">일일 검토</h1>
      <div class="profile-tabs">
        <a class="profile-tab {'active' if user_profile == 'hq' else ''}" href="/daily?user_profile=hq&period_mode={escape(period_mode)}&window_days={window_days}&review_limit={review_limit}&status_filter={escape(status_filter)}">본사</a>
        <a class="profile-tab {'active' if user_profile == 'jaehwa' else ''}" href="/daily?user_profile=jaehwa&period_mode={escape(period_mode)}&window_days={window_days}&review_limit={review_limit}&status_filter={escape(status_filter)}">재화</a>
      </div>
      <form method="get" class="toolbar">
        <input type="hidden" name="user_profile" value="{escape(user_profile)}">
        <div>
          <label for="period_mode">기간 기준</label>
          <select id="period_mode" name="period_mode">
            <option value="previous_day" {'selected' if period_mode == 'previous_day' else ''}>전일 마감 기준</option>
            <option value="last24h" {'selected' if period_mode == 'last24h' else ''}>현재 시점 최근 24h</option>
          </select>
        </div>
        <div>
          <label for="window_days">기간(일)</label>
          <input id="window_days" name="window_days" type="number" min="1" max="30" value="{window_days}">
        </div>
        <div>
          <label for="review_limit">조회 건수 상한</label>
          <input id="review_limit" name="review_limit" type="number" min="1" max="1000" value="{review_limit}">
        </div>
        <div>
          <label for="status_filter">상태 필터</label>
          <select id="status_filter" name="status_filter">
            <option value="all" {'selected' if status_filter == 'all' else ''}>전체</option>
            <option value="pending" {'selected' if status_filter == 'pending' else ''}>미검토</option>
            <option value="reviewed" {'selected' if status_filter == 'reviewed' else ''}>검토완료</option>
            <option value="ignored" {'selected' if status_filter == 'ignored' else ''}>제외</option>
          </select>
        </div>
        <a class="toolbar-link" href="{escape(daily_export_href)}">엑셀 다운로드</a>
        <button type="submit">새로고침</button>
      </form>
      <div class="quick-filters">
        {quick_period_links}
      </div>
      <div class="quick-filters">
        {quick_status_links}
      </div>
      <form id="bulk-form" method="post" class="toolbar bulk-toolbar">
        <input type="hidden" name="user_profile" value="{escape(user_profile)}">
        <input type="hidden" name="period_mode" value="{escape(period_mode)}">
        <input type="hidden" name="window_days" value="{window_days}">
        <input type="hidden" name="review_limit" value="{review_limit}">
        <input type="hidden" name="status_filter" value="{escape(status_filter)}">
        <div>
          <label for="bulk_status">일괄 상태</label>
          <select id="bulk_status" name="bulk_status">
            <option value="pending">미검토</option>
            <option value="reviewed">검토완료</option>
            <option value="ignored">제외</option>
          </select>
        </div>
        <div style="min-width:280px;">
          <label for="bulk_note">일괄 메모(선택)</label>
          <input id="bulk_note" name="bulk_note" placeholder="선택된 CVE에 동일 메모 저장">
        </div>
        <button type="submit" name="action" value="bulk_update">선택 항목 일괄 저장</button>
        <button type="submit" name="action" value="undo_bulk" {'disabled' if not has_undo_bulk else ''}>최근 일괄 변경 되돌리기</button>
      </form>
      {notice_html}
      {error_html}
      <p class="meta">대상 {total_count}건 (표시 {filtered_count}건) | {' | '.join(escape(line) for line in info_lines)}</p>
      <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th class="sticky-col sticky-left"><input id="bulk-select-all" type="checkbox" title="전체 선택"></th><th>CVE ID</th><th>CVSS</th><th>Type</th><th>Last Modified</th><th>Description</th><th>Preset</th><th class="sticky-col sticky-right">Review</th>
          </tr>
        </thead>
        <tbody>
          {''.join(row_chunks)}
        </tbody>
      </table>
      </div>
    </section>
  </main>
</body>
<script>
  (() => {{
    const selectAll = document.querySelector("#bulk-select-all");
    const checks = () => Array.from(document.querySelectorAll(".bulk-cve-check"));
    selectAll?.addEventListener("change", () => {{
      const checked = !!selectAll.checked;
      checks().forEach((el) => {{
        el.checked = checked;
      }});
    }});
    checks().forEach((el) => {{
      el.addEventListener("change", () => {{
        const all = checks();
        if (!all.length || !selectAll) return;
        selectAll.checked = all.every((x) => x.checked);
      }});
    }});
    const bulkForm = document.querySelector("#bulk-form");
    bulkForm?.addEventListener("submit", (event) => {{
      const submitter = event.submitter;
      if (!submitter || submitter.value !== "bulk_update") {{
        return;
      }}
      const selected = checks().filter((x) => x.checked).length;
      if (selected <= 0) {{
        event.preventDefault();
        window.alert("일괄 변경할 CVE를 선택하세요.");
        return;
      }}
      const targetStatus = (document.querySelector("#bulk_status") || {{ value: "pending" }}).value;
      const confirmed = window.confirm(`선택한 ${{selected}}건을 '${{targetStatus}}' 상태로 변경할까요?`);
      if (!confirmed) {{
        event.preventDefault();
      }}
    }});
    if ({'true' if bool(notice_text) else 'false'}) {{
      const toast = document.createElement("div");
      toast.textContent = "{escape(notice_text)}";
      toast.style.position = "fixed";
      toast.style.right = "18px";
      toast.style.bottom = "18px";
      toast.style.padding = "10px 14px";
      toast.style.border = "1px solid #9ecfb9";
      toast.style.background = "#e7f6ef";
      toast.style.color = "#0e6a4c";
      toast.style.borderRadius = "10px";
      toast.style.fontSize = "12px";
      toast.style.fontWeight = "700";
      toast.style.zIndex = "999";
      document.body.appendChild(toast);
      setTimeout(() => {{
        toast.remove();
      }}, 1500);
    }}
  }})();
</script>
</html>
"""


@app.get("/daily_export.xlsx")
def daily_export_xlsx() -> object:
    user_profile = _normalize_user_profile(request.args.get("user_profile"))
    period_mode = (request.args.get("period_mode") or "previous_day").strip().lower()
    if period_mode not in {"previous_day", "last24h"}:
        period_mode = "previous_day"
    status_filter = (request.args.get("status_filter") or "pending").strip().lower()
    if status_filter not in {"all", "pending", "reviewed", "ignored"}:
        status_filter = "pending"

    try:
        settings_obj = load_settings(".env")
        profile_defaults = fetch_profile_settings(settings_obj, user_profile)
    except Exception as exc:  # pragma: no cover
        return f"Failed to load settings: {exc}", 500

    now_local = datetime.now().replace(second=0, microsecond=0)
    window_days_raw = request.args.get("window_days") or str(profile_defaults["daily_review_window_days"])
    try:
        window_days = max(1, min(int(window_days_raw), 30))
    except ValueError:
        window_days = int(profile_defaults["daily_review_window_days"])
    review_limit_raw = request.args.get("review_limit") or str(profile_defaults["daily_review_limit"])
    try:
        review_limit = max(1, min(int(review_limit_raw), 1000))
    except ValueError:
        review_limit = int(profile_defaults["daily_review_limit"])

    if period_mode == "previous_day":
        end_dt = now_local.replace(hour=0, minute=0)
        start_dt = end_dt - timedelta(days=window_days)
        review_date = (end_dt - timedelta(days=1)).date().isoformat()
    else:
        end_dt = now_local
        start_dt = now_local - timedelta(hours=24 * window_days)
        review_date = end_dt.date().isoformat()

    try:
        active_presets = [item for item in fetch_profile_presets(settings_obj, user_profile) if item["is_enabled"]]
    except Exception:
        active_presets = []

    matched_preset_map: dict[str, list[str]] = {}
    rows: list[dict[str, object]] = []
    if active_presets:
        merged_by_cve: dict[str, dict[str, object]] = {}
        for preset in active_presets:
            preset_name = str(preset["preset_name"])
            rule = dict(preset["rule"])
            preset_rows, _ = fetch_cves_from_db(
                settings_obj,
                str(rule["product"]) or None,
                str(rule["vendor"]) or None,
                str(rule["keyword"]) or None,
                list(rule["impact_type"]) or None,
                float(rule["min_cvss"]),
                review_limit,
                offset=0,
                sort_by="last_modified",
                sort_order="desc",
                last_modified_start=start_dt,
                last_modified_end=end_dt,
                cpe_missing_only=bool(rule["cpe_missing_only"]),
                cpe_objects=list(rule["cpe_objects_catalog"]) or None,
                include_total_count=False,
            )
            for row in preset_rows:
                cve_id = str(row.get("id", ""))
                if not cve_id:
                    continue
                if cve_id not in merged_by_cve:
                    merged_by_cve[cve_id] = row
                matched_preset_map.setdefault(cve_id, [])
                if preset_name not in matched_preset_map[cve_id]:
                    matched_preset_map[cve_id].append(preset_name)
        rows = list(merged_by_cve.values())
        rows.sort(
            key=lambda row: (
                row.get("last_modified_at") or datetime.min,
                float(row.get("cvss_score") or 0.0),
            ),
            reverse=True,
        )
        rows = rows[:review_limit]
    else:
        rows, _ = fetch_cves_from_db(
            settings_obj,
            str(profile_defaults["product"]) or None,
            str(profile_defaults["vendor"]) or None,
            str(profile_defaults["keyword"]) or None,
            list(profile_defaults["impact_type"]) or None,
            float(profile_defaults["min_cvss"]),
            review_limit,
            offset=0,
            sort_by="last_modified",
            sort_order="desc",
            last_modified_start=start_dt,
            last_modified_end=end_dt,
            cpe_missing_only=bool(profile_defaults["cpe_missing_only"]),
            cpe_objects=list(profile_defaults["cpe_objects_catalog"]) or None,
            include_total_count=False,
        )

    review_map = fetch_daily_review_map(settings_obj, user_profile, review_date)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Review"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append(["Generated At", generated_at])
    sheet.append(["Profile", user_profile])
    sheet.append(["Period Mode", period_mode])
    sheet.append(["Window Days", str(window_days)])
    sheet.append(["Status Filter", status_filter])
    sheet.append([])
    headers = ["CVE ID", "CVSS", "Type", "Last Modified", "Description", "Preset", "Review Status", "Review Note"]
    sheet.append(headers)

    header_row = sheet.max_row
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="2F3F45")
        cell.fill = PatternFill(fill_type="solid", start_color="F2EEE4", end_color="F2EEE4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        cve_id = str(row.get("id", ""))
        state = review_map.get(cve_id, {"status": "pending", "note": ""})
        row_status = str(state.get("status", "pending"))
        if status_filter != "all" and row_status != status_filter:
            continue
        sheet.append(
            [
                cve_id,
                str(row.get("cvss_score") if row.get("cvss_score") is not None else "0.0"),
                str(row.get("vuln_type", "Other")),
                format_last_modified(row.get("last_modified_at", "N/A")),
                str(row.get("description", "")),
                ", ".join(matched_preset_map.get(cve_id, [])),
                row_status,
                str(state.get("note", "")),
            ]
        )

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 90
    sheet.column_dimensions["F"].width = 34
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 42

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=5).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_idx, column=8).alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"daily_review_{user_profile}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Run local web UI for CVE query")
    parser.add_argument("--host", default="0.0.0.0", help="Bind host")
    parser.add_argument("--port", type=int, default=8888, help="Bind port")
    args = parser.parse_args()
    app.run(host=args.host, port=args.port, debug=False)


if __name__ == "__main__":
    main()
